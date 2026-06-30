"""run_swarm surfaces deterministic detection findings (no silent corruption)."""
import openpyxl
from mcg_swarm.runner import run_swarm

def _save(tmp_path, name, build):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Data"; build(ws)
    p = tmp_path / name; wb.save(p); return str(p)

def test_stacked_tables_flag_uncovered_data(tmp_path):
    def build(ws):
        ws.append(["Region", "Rev", "Units"]); ws.append(["NA", 1, 2])
        ws.append([]); ws.append([])
        ws.append(["Product", "Price", "SKU"]); ws.append(["Widget", 9, "W1"])
    ext = run_swarm(_save(tmp_path, "stacked.xlsx", build))
    cats = [f.category for f in ext.findings]
    assert "uncovered-data" in cats
    assert ext.errors  # derived, non-empty — not silent

def test_side_by_side_flag_uncovered_data(tmp_path):
    def build(ws):
        ws["A1"], ws["B1"] = "Region", "Rev"
        ws["D1"], ws["E1"] = "Product", "Price"
        ws["A2"], ws["B2"] = "NA", 1
        ws["D2"], ws["E2"] = "Widget", 9
    ext = run_swarm(_save(tmp_path, "sbs.xlsx", build))
    assert "uncovered-data" in [f.category for f in ext.findings]

def test_empty_corner_flagged_on_table(tmp_path):
    def build(ws):
        ws.append(["", "Q1", "Q2"]); ws.append(["Revenue", 1, 2]); ws.append(["COGS", 3, 4])
    ext = run_swarm(_save(tmp_path, "transposed.xlsx", build))
    all_cats = [f.category for t in ext.tables for f in t.findings] + \
               [f.category for f in ext.findings]
    assert "empty-header-corner" in all_cats

def test_clean_workbook_no_detection_findings(tmp_path):
    def build(ws):
        ws.append(["Region", "Rev"]); ws.append(["NA", 1]); ws.append(["EU", 2])
    ext = run_swarm(_save(tmp_path, "clean.xlsx", build))
    detection = {"uncovered-data", "empty-header-corner", "false-header-span",
                 "transpose-suspected"}
    found = {f.category for f in ext.findings} | \
            {f.category for t in ext.tables for f in t.findings}
    assert not (found & detection)
    assert ext.errors == []
