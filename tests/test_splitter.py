import openpyxl, pytest
from mcg_swarm.splitter import split_workbook, TableHandle, detect_table

def _wb(tmp_path, rows, name="t.xlsx"):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Data"
    for r in rows: ws.append(r)
    p = tmp_path / name; wb.save(p); return str(p)

def test_clean_table_handle(tmp_path):
    p = _wb(tmp_path, [["Region", "Revenue", "Units"],
                       ["EMEA", 100, 5], ["APAC", 200, 9]])
    handles = split_workbook(p)
    assert len(handles) == 1
    h = handles[0]
    assert h.sheet == "Data" and h.header_row == 1 and not h.ambiguous
    assert h.region == "A1:C3"
    assert [c.name for c in h.columns] == ["Region", "Revenue", "Units"]
    assert h.columns[0].role == "key"
    assert h.columns[1].dtype == "number" and h.columns[0].dtype == "string"

def test_title_banner_offset_is_detected_not_guessed(tmp_path):
    p = _wb(tmp_path, [["Q3 Sales Report"], [],
                       ["Region", "Revenue"], ["EMEA", 100]])
    h = split_workbook(p)[0]
    # deterministic detector may resolve the offset; if it can't, it must flag, not guess
    assert h.header_row == 3 or h.ambiguous

def test_empty_sheet_is_ambiguous(tmp_path):
    p = _wb(tmp_path, [])
    h = split_workbook(p)[0]
    assert h.ambiguous


def test_merged_title_banner_above_header(tmp_path):
    """Row 1 has ONE non-empty cell (title banner), row 2 is the real multi-column header."""
    # This mimics a merged title cell spanning A1:C1 — openpyxl reads the rest as None
    p = _wb(tmp_path, [
        ["Title", None, None],        # row 1: single-cell title banner (like merged A1:C1)
        ["Region", "Q1", "Q2"],       # row 2: real header
        ["North", 100, 200],          # row 3: data
        ["South", 300, 400],          # row 4: data
    ])
    h = split_workbook(p)[0]
    assert h.header_row == 2, f"Expected header_row=2, got {h.header_row}"
    assert not h.ambiguous, f"Should not be ambiguous, got reason={h.reason!r}"
    assert [c.name for c in h.columns] == ["Region", "Q1", "Q2"], \
        f"Expected 3 columns, got {[c.name for c in h.columns]}"
    # Pattern A: region NOW includes the banner row above the header, so region
    # starts at A1 (not A2). header_row stays 2 (the actual header location).
    assert h.region.startswith("A1"), \
        f"Region should start at A1 (banner included), got {h.region!r}"


def test_left_offset_table_trims_empty_leading_column(tmp_path):
    """Column A is entirely empty; table lives in B:C starting row 2."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    # Row 1: entirely empty
    ws.append([None, None, None])
    # Row 2: header in B:C (A is empty)
    ws.cell(row=2, column=2, value="Name")
    ws.cell(row=2, column=3, value="Val")
    # Row 3: data
    ws.cell(row=3, column=2, value="Alice")
    ws.cell(row=3, column=3, value=42)
    # Row 4: data
    ws.cell(row=4, column=2, value="Bob")
    ws.cell(row=4, column=3, value=99)
    p = tmp_path / "offset.xlsx"
    wb.save(p)
    h = split_workbook(str(p))[0]
    assert not h.ambiguous, f"Should not be ambiguous, got reason={h.reason!r}"
    assert h.region.startswith("B"), f"Region should start at B, got {h.region!r}"
    col_names = [c.name for c in h.columns]
    assert col_names == ["Name", "Val"], f"Expected ['Name','Val'], got {col_names}"
    assert "A" not in col_names, f"Phantom 'A' column should not appear"
    assert h.columns[0].role == "key", "First real column should be key"


# ── Pattern A: region top includes contiguous banner rows above header ─────────

def test_region_includes_title_banner_above_header(tmp_path):
    """PATTERN A: single title row immediately above header → region top == banner row."""
    p = _wb(tmp_path, [
        ["Report", None],          # row 1: title banner
        ["Region", "Rev"],         # row 2: real header
        ["East", 500],             # row 3: data
        ["West", 300],             # row 4: data
    ])
    h = split_workbook(p)[0]
    assert not h.ambiguous, f"Should not be ambiguous: {h.reason!r}"
    assert h.header_row == 2, f"header_row should be 2, got {h.header_row}"
    # Region top must be row 1 (includes the banner)
    assert h.region.startswith("A1"), f"Region should start at A1, got {h.region!r}"
    assert h.region.endswith("B4"), f"Region should end at B4, got {h.region!r}"
    assert [c.name for c in h.columns] == ["Region", "Rev"]


# ── Pattern B: right edge stops at first fully-empty gap column ────────────────

def test_right_edge_stops_at_gap_column(tmp_path):
    """PATTERN B: empty col C separates real table (A:B) from stray params (D)."""
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Data"
    # Row 1: header A=Name, B=Val, C=empty, D=StrayParam
    ws.cell(row=1, column=1, value="Name")
    ws.cell(row=1, column=2, value="Val")
    ws.cell(row=1, column=3, value=None)      # gap column
    ws.cell(row=1, column=4, value="StrayParam")
    # Row 2: data only in A:B
    ws.cell(row=2, column=1, value="Alpha")
    ws.cell(row=2, column=2, value=10)
    # Row 3: stray value in D only (no A:B data; but A:B have content in rows above)
    ws.cell(row=3, column=4, value=42)
    p = tmp_path / "gap.xlsx"; wb.save(p)
    h = split_workbook(str(p))[0]
    assert not h.ambiguous, f"Should not be ambiguous: {h.reason!r}"
    # Region must stop at column B
    assert "D" not in h.region, f"Stray col D should be excluded, got {h.region!r}"
    assert h.region.endswith("B2") or h.region.endswith("B3") or "B" in h.region.split(":")[1], \
        f"Region right edge should be B, got {h.region!r}"
    col_names = [c.name for c in h.columns]
    assert "StrayParam" not in col_names, f"StrayParam should be excluded, got {col_names}"
    assert col_names == ["Name", "Val"]
