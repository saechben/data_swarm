"""Builders for adversarial workbooks (committed stress fixtures)."""
import openpyxl


def _wb():
    wb = openpyxl.Workbook(); wb.active.title = "Data"; return wb, wb.active


def two_stacked(path):
    wb, ws = _wb()
    ws.append(["Region", "Revenue", "Units"]); ws.append(["NA", 100, 5]); ws.append(["EU", 200, 9])
    ws.append([]); ws.append([])
    ws.append(["Product", "Price", "SKU"]); ws.append(["Widget", 9.99, "W-1"])
    wb.save(path); return str(path)


def side_by_side(path):
    wb, ws = _wb()
    ws["A1"], ws["B1"], ws["C1"] = "Region", "Revenue", "Units"
    ws["E1"], ws["F1"], ws["G1"] = "Product", "Price", "Stock"
    ws["A2"], ws["B2"], ws["C2"] = "NA", 100, 5
    ws["E2"], ws["F2"], ws["G2"] = "Widget", 9.99, 12
    wb.save(path); return str(path)


def preamble_rows(path):
    wb, ws = _wb()
    ws.append(["Report", "generated 2024-06-01", "by system"])
    ws.append(["Confidential", "do not distribute", ""])
    ws.append([])
    ws.append(["Region", "Revenue", "Units"]); ws.append(["NA", 100, 5]); ws.append(["EU", 200, 9])
    wb.save(path); return str(path)


def transposed(path):
    wb, ws = _wb()
    ws.append(["", "Q1", "Q2", "Q3"]); ws.append(["Revenue", 100, 120, 130]); ws.append(["COGS", 40, 48, 52])
    wb.save(path); return str(path)
