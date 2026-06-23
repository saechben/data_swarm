# tests/test_extraction.py
import openpyxl, pytest
from mcg_swarm.splitter import split_workbook
from mcg_swarm.extraction import build_index

def _wb(tmp_path, rows):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Data"
    for r in rows: ws.append(r)
    p = tmp_path / "t.xlsx"; wb.save(p); return str(p)

def test_query_by_key_and_column(tmp_path):
    p = _wb(tmp_path, [["Region", "Revenue"], ["EMEA", 100], ["APAC", 200]])
    h = split_workbook(p)[0]
    idx = build_index(p, h, row_key=["Region"])
    v = idx.query("APAC", "Revenue")
    assert v.value == 200 and v.cell_ref == "B3" and v.sheet == "Data" and v.dtype == "number"

def test_unknown_key_and_column_raise(tmp_path):
    p = _wb(tmp_path, [["Region", "Revenue"], ["EMEA", 100]])
    idx = build_index(p, split_workbook(p)[0], row_key=["Region"])
    with pytest.raises(KeyError): idx.query("NOPE", "Revenue")
    with pytest.raises(KeyError): idx.query("EMEA", "NoCol")

def test_live_read_reflects_edits(tmp_path):
    p = _wb(tmp_path, [["Region", "Revenue"], ["EMEA", 100]])
    idx = build_index(p, split_workbook(p)[0], row_key=["Region"])
    assert idx.query("EMEA", "Revenue").value == 100
    wb = openpyxl.load_workbook(p); wb["Data"]["B2"] = 999; wb.save(p)
    assert idx.query("EMEA", "Revenue").value == 999  # no rebuild

def test_query_range_reads_all_cells(tmp_path):
    # 3x2 table: header + 2 data rows, 2 value columns
    p = _wb(tmp_path, [["X", "Y"], [10, 20], [30, 40], [50, 60]])
    idx = build_index(p, split_workbook(p)[0], row_key=["X"])
    # query_range over A2:B4 (the 3 data rows, both columns)
    results = idx.query_range("A2:B4")
    assert len(results) == 6
    by_ref = {v.cell_ref: v.value for v in results}
    assert by_ref["A2"] == 10
    assert by_ref["B2"] == 20
    assert by_ref["A3"] == 30
    assert by_ref["B3"] == 40
    assert by_ref["A4"] == 50
    assert by_ref["B4"] == 60
    # all provenance fields set correctly
    assert all(v.sheet == "Data" for v in results)
    assert all(v.dtype == "number" for v in results)
    assert all(v.unit is None for v in results)
    assert all(v.is_computed is False for v in results)


def test_coverage_helpers(tmp_path):
    p = _wb(tmp_path, [["Region", "Revenue"], ["EMEA", 100], ["APAC", 200]])
    idx = build_index(p, split_workbook(p)[0], row_key=["Region"])
    assert set(idx.row_keys()) == {"EMEA", "APAC"}
    assert idx.column_names() == ["Region", "Revenue"]


# --- read_all tests ---

def test_read_all_matches_query(tmp_path):
    """read_all returns same values as individual query() calls."""
    p = _wb(tmp_path, [["Region", "Revenue", "Units"],
                        ["EMEA", 100, 5],
                        ["APAC", 200, 10],
                        ["AMER", 300, 15]])
    idx = build_index(p, split_workbook(p)[0], row_key=["Region"])
    rows = idx.read_all()
    # Should have 3 row_keys × 3 columns = 9 tuples
    assert len(rows) == 9
    # Build lookup from read_all
    lookup = {(rk, cn): (val, ref) for rk, cn, val, ref in rows}
    # Compare to query() for all combinations
    for row_key in ["EMEA", "APAC", "AMER"]:
        for col_name in ["Region", "Revenue", "Units"]:
            qval = idx.query(row_key, col_name)
            ra_val, ra_ref = lookup[(row_key, col_name)]
            assert ra_val == qval.value, f"mismatch at ({row_key!r}, {col_name!r})"
            assert ra_ref == qval.cell_ref, f"cell_ref mismatch at ({row_key!r}, {col_name!r})"


def test_read_all_cell_refs_correct(tmp_path):
    """read_all returns correct A1-style cell refs."""
    p = _wb(tmp_path, [["Region", "Revenue"], ["EMEA", 100], ["APAC", 200]])
    idx = build_index(p, split_workbook(p)[0], row_key=["Region"])
    rows = idx.read_all()
    by_ref = {ref: (rk, cn, val) for rk, cn, val, ref in rows}
    # header is row 1; data starts row 2
    assert "A2" in by_ref  # EMEA, Region column
    assert by_ref["B2"][2] == 100   # Revenue for EMEA
    assert by_ref["B3"][2] == 200   # Revenue for APAC


def test_read_all_max_rows_bounds(tmp_path):
    """max_rows caps the number of row keys returned."""
    p = _wb(tmp_path, [["Region", "Revenue"],
                        ["R1", 1], ["R2", 2], ["R3", 3], ["R4", 4], ["R5", 5]])
    idx = build_index(p, split_workbook(p)[0], row_key=["Region"])
    # Without cap: 5 row_keys × 2 columns = 10
    assert len(idx.read_all()) == 10
    # With max_rows=3: only first 3 row_keys → 3×2 = 6
    capped = idx.read_all(max_rows=3)
    assert len(capped) == 6
    row_keys_seen = {rk for rk, _, _, _ in capped}
    assert row_keys_seen == {"R1", "R2", "R3"}


# ── Pattern A: extraction works when region top is ABOVE header row ────────────

def test_index_resolves_when_region_top_above_header(tmp_path):
    """PATTERN A: build_index on a handle where region includes a banner row above header."""
    p = _wb(tmp_path, [
        ["Report", None],          # row 1: title banner (part of region)
        ["Region", "Rev"],         # row 2: real header
        ["East", 500],             # row 3: data
        ["West", 300],             # row 4: data
    ])
    h = split_workbook(p)[0]
    assert h.header_row == 2, f"Pre-condition: header_row must be 2, got {h.header_row}"
    assert h.region.startswith("A1"), f"Pre-condition: region must start A1, got {h.region}"
    idx = build_index(p, h, row_key=["Region"])
    # East → Rev should be 500 at cell B3
    v = idx.query("East", "Rev")
    assert v.value == 500, f"Expected 500, got {v.value}"
    assert v.cell_ref == "B3", f"Expected B3, got {v.cell_ref}"
    # West → Rev should be 300 at cell B4
    v2 = idx.query("West", "Rev")
    assert v2.value == 300, f"Expected 300, got {v2.value}"
    assert v2.cell_ref == "B4", f"Expected B4, got {v2.cell_ref}"


# ── Pattern C: ExtractionIndex with 2-row composite header ────────────────────

def test_index_with_two_row_header(tmp_path):
    """Pattern C: build_index on a 2-row-header handle resolves composite column names."""
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Data"
    ws.append([None, "Product", "Revenue", None,  None])      # row1: group header
    ws.append([None, None,      "Gross",   "Net", "Margin"])   # row2: leaf header
    ws.append([None, "Product A", 1000, 850, 85])               # row3: data
    ws.append([None, "Product B", 1200, 900, 75])               # row4: data
    p = tmp_path / "tworow.xlsx"; wb.save(p)
    h = split_workbook(str(p))[0]
    assert h.header_span == 2
    idx = build_index(str(p), h, row_key=["Product"])
    # Query a leaf-col value
    v = idx.query("Product A", "Gross")
    assert v.value == 1000, f"Expected 1000, got {v.value}"
    assert v.cell_ref == "C3", f"Expected C3, got {v.cell_ref}"
    # Query Net for Product B
    v2 = idx.query("Product B", "Net")
    assert v2.value == 900, f"Expected 900, got {v2.value}"
    assert v2.cell_ref == "D4", f"Expected D4, got {v2.cell_ref}"
    # Query the key column itself
    v3 = idx.query("Product A", "Product")
    assert v3.value == "Product A"
    assert v3.cell_ref == "B3"


def test_read_all_none_values_preserved(tmp_path):
    """read_all returns None for empty cells (no silently dropping)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Region", "Revenue"])
    ws.append(["EMEA", None])  # empty revenue
    ws.append(["APAC", 200])
    p = tmp_path / "t.xlsx"
    wb.save(str(p))
    idx = build_index(str(p), split_workbook(str(p))[0], row_key=["Region"])
    rows = idx.read_all()
    by_key_col = {(rk, cn): val for rk, cn, val, _ in rows}
    assert by_key_col[("EMEA", "Revenue")] is None
    assert by_key_col[("APAC", "Revenue")] == 200
