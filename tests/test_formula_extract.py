import openpyxl
from mcg_swarm.source import OpenpyxlFileSource


def _write_vertical_formula_wb(path):
    """3-col table: Units | Price | Revenue(=A*B per row). Header row 1, data rows 2-4."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Units", "Price", "Revenue"])
    for r in range(2, 5):
        ws.cell(row=r, column=1, value=r)           # Units
        ws.cell(row=r, column=2, value=10)          # Price
        ws.cell(row=r, column=3, value=f"=A{r}*B{r}")  # Revenue formula
    wb.save(path)


def test_read_formula_region_returns_formula_strings(tmp_path):
    p = tmp_path / "vf.xlsx"
    _write_vertical_formula_wb(str(p))
    src = OpenpyxlFileSource(str(p))
    rows = src.read_formula_region("Sheet1", 2, 1, 4, 3)
    # row 2 (first data row): Units=2, Price=10, Revenue="=A2*B2"
    assert rows[0][2] == "=A2*B2"
    assert rows[2][2] == "=A4*B4"


def test_read_formula_region_empty_cells_no_crash(tmp_path):
    """Regression for the old EmptyCell cell.coordinate crash (commit b77195b)."""
    p = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B", "C"])
    ws.append([1, None, None])   # row 2 has empty cells
    wb.save(str(p))
    src = OpenpyxlFileSource(str(p))
    rows = src.read_formula_region("Sheet1", 1, 1, 2, 3)  # must not raise
    assert rows[1][0] == 1 and rows[1][1] is None


from fake_source import FakeSource, vertical_fake


def test_index_geometry_accessors():
    from mcg_swarm.splitter import split_workbook
    from mcg_swarm.extraction import build_index
    src = vertical_fake()
    handle = split_workbook(src)[0]
    index = build_index(src, handle, row_key=[])
    cols = index.physical_columns()
    assert cols["Units"] == 1 and cols["Price"] == 2 and cols["Revenue"] == 3
    assert index.data_row_numbers() == [2, 3, 4]


def _make_index(src, row_key=None):
    from mcg_swarm.splitter import split_workbook
    from mcg_swarm.extraction import build_index
    handle = split_workbook(src)[0]
    index = build_index(src, handle, row_key=row_key or [])
    return handle, index


def test_extract_vertical_formula_translates_and_upgrades_role():
    from mcg_swarm.formula_extract import extract_formulas
    src = vertical_fake()
    handle, index = _make_index(src)
    columns = list(handle.columns)
    formulas, notes = extract_formulas(src, index, columns)
    rev = next(f for f in formulas if f.target == "Revenue")
    assert rev.expression == "Units*Price"
    assert {o.name for o in rev.operands} == {"Units", "Price"}
    assert rev.context and "Revenue" in rev.context
    # role upgraded in place on the passed ColumnSpec
    assert next(c for c in columns if c.name == "Revenue").role == "computed"


def test_extract_cross_sheet_captured_untranslated():
    from mcg_swarm.formula_extract import extract_formulas
    values = {(1, 1): "A", (1, 2): "B", (1, 3): "C"}
    formulas_in = {}
    for r in range(2, 5):
        values[(r, 1)] = r
        values[(r, 2)] = 2
        values[(r, 3)] = 99               # arbitrary cached value
        formulas_in[(r, 3)] = f"='Inputs'!A1*A{r}"   # cross-sheet -> untranslatable
    src = FakeSource("Sheet1", values, formulas_in)
    handle, index = _make_index(src)
    columns = list(handle.columns)
    formulas, notes = extract_formulas(src, index, columns)
    c = next(f for f in formulas if f.target == "C")
    assert c.operands == []                       # not translated
    assert c.context                              # reason present
    assert next(col for col in columns if col.name == "C").role != "computed"
    assert any("C" in n for n in notes)           # provisional note recorded


def test_extract_no_formulas_returns_empty():
    from mcg_swarm.formula_extract import extract_formulas
    values = {(1, 1): "A", (1, 2): "B"}
    for r in range(2, 5):
        values[(r, 1)] = r
        values[(r, 2)] = r * 2
    src = FakeSource("Sheet1", values, {})
    handle, index = _make_index(src)
    formulas, notes = extract_formulas(src, index, list(handle.columns))
    assert formulas == [] and notes == []


def test_orchestrator_populates_formulas_endtoend():
    from mcg_swarm.splitter import split_workbook
    from mcg_swarm.orchestrator import _orchestrate_core
    src = vertical_fake()
    handle = split_workbook(src)[0]
    table = _orchestrate_core(src, handle, table_id="t0")
    rev = [f for f in table.formulas if f.target == "Revenue"]
    assert rev and rev[0].expression == "Units*Price"
    assert rev[0].context
    assert table.errors == []   # gate stays green (formula recomputes correctly)


def test_per_formula_isolation_bad_formula_does_not_wipe_good():
    """Critical regression: one untranslatable formula in a column must NOT abort
    extraction of other columns in the same table. The translatable column must
    still appear as computed, and the untranslatable column as captured-untranslated,
    and NO whole-table 'formula extraction error' note may be emitted."""
    from mcg_swarm.formula_extract import extract_formulas

    # 4-col table: Units | Price | Revenue (=A*B, translatable) | Flag (=A>B, untranslatable)
    values = {
        (1, 1): "Units", (1, 2): "Price", (1, 3): "Revenue", (1, 4): "Flag"
    }
    formulas_in = {}
    for r in range(2, 5):
        values[(r, 1)] = r
        values[(r, 2)] = 10
        values[(r, 3)] = r * 10          # cached value for Revenue
        values[(r, 4)] = 0               # cached value for Flag
        formulas_in[(r, 3)] = f"=A{r}*B{r}"   # translatable
        formulas_in[(r, 4)] = f"=A{r}>B{r}"   # untranslatable comparison

    src = FakeSource("Sheet1", values, formulas_in)
    handle, index = _make_index(src)
    columns = list(handle.columns)
    formulas, notes = extract_formulas(src, index, columns)

    # No whole-table abort
    assert not any("formula extraction error" in n for n in notes), \
        f"extraction aborted with error: {notes}"

    # Revenue is translated (computed role, has operands)
    revenue = next((f for f in formulas if f.target == "Revenue"), None)
    assert revenue is not None, "Revenue formula not found"
    assert revenue.expression == "Units*Price"
    assert len(revenue.operands) > 0
    rev_col = next(c for c in columns if c.name == "Revenue")
    assert rev_col.role == "computed"

    # Flag is captured-untranslated (empty operands, reason present)
    flag = next((f for f in formulas if f.target == "Flag"), None)
    assert flag is not None, "Flag formula not captured"
    assert flag.operands == []
    flag_col = next(c for c in columns if c.name == "Flag")
    assert flag_col.role != "computed"


def test_real_transposed_workbook_captures_without_crash():
    """formula_chain_pnl is transposed (=B2*B3). Phase 1 captures these as
    untranslated (same-row guard) and must never crash or mark them computed."""
    import os
    wb_path = os.path.join("eval", "data", "workbooks", "formula_chain_pnl.xlsx")
    if not os.path.exists(wb_path):
        import pytest
        pytest.skip("extreme workbook not generated")
    from mcg_swarm.runner import run_swarm
    ext = run_swarm({"main": wb_path})
    # never raises; some table carries captured (untranslated) formulas or notes
    assert ext.tables  # extraction produced tables
    for t in ext.tables:
        for f in t.formulas:
            # any formula present is either translated (has operands) or captured
            assert f.context is not None
