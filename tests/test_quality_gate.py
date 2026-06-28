# tests/test_testing.py
import openpyxl
from mcg_swarm.quality_gate import run_table_tests, TableTestReport
from mcg_swarm.source import OpenpyxlFileSource
from mcg_swarm.splitter import split_workbook
from mcg_swarm.extraction import build_index
from mcg_swarm.schemas import CanonicalTable, ExtractionRef


def _wb(tmp_path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for r in rows:
        ws.append(r)
    p = tmp_path / "t.xlsx"
    wb.save(p)
    return str(p)


def _canon(h):
    return CanonicalTable(
        table_id="t",
        sheet=h.sheet,
        region=h.region,
        header_row=h.header_row,
        columns=h.columns,
        description="d",
        extraction=ExtractionRef(script_name="idx", row_key=[h.columns[0].name]),
    )


def test_passes_on_clean_table(tmp_path):
    p = _wb(tmp_path, [["Region", "Revenue"], ["EMEA", 100], ["APAC", 200]])
    h = split_workbook(p)[0]
    idx = build_index(p, h, row_key=["Region"])
    rep = run_table_tests(OpenpyxlFileSource(p), _canon(h), idx)
    assert rep.passed and rep.failures == []


def test_detects_roundtrip_mismatch(tmp_path):
    """Column-integrity (and round-trip) must catch a string-col-into-number-col remap."""
    p = _wb(tmp_path, [["Region", "Revenue"], ["EMEA", 100]])
    h = split_workbook(p)[0]
    idx = build_index(p, h, row_key=["Region"])
    # corrupt: Revenue now points to Region's physical column
    idx._col_to_phys["Revenue"] = idx._col_to_phys["Region"]
    rep = run_table_tests(OpenpyxlFileSource(p), _canon(h), idx)
    # column-integrity catches it: Revenue col in index != Revenue col in live header
    assert not rep.passed and rep.failures


def test_detects_numeric_column_swap(tmp_path):
    """Regression guard: numeric->numeric column remap must be caught by column-integrity.

    A Revenue/Units swap produces self-consistent cell_ref+value in the index,
    so the round-trip alone cannot detect it.  The column-integrity check reads
    the live header independently and must flag the mismatch.
    """
    p = _wb(tmp_path, [["Region", "Revenue", "Units"], ["EMEA", 100, 5], ["APAC", 200, 8]])
    h = split_workbook(p)[0]
    idx = build_index(p, h, row_key=["Region"])
    # swap Revenue and Units — both numeric so round-trip cannot distinguish them
    idx._col_to_phys["Revenue"], idx._col_to_phys["Units"] = (
        idx._col_to_phys["Units"],
        idx._col_to_phys["Revenue"],
    )
    rep = run_table_tests(OpenpyxlFileSource(p), _canon(h), idx)
    assert not rep.passed, f"Expected failure but got no failures: {rep.failures}"
    assert any("column-integrity" in f for f in rep.failures), (
        f"Expected column-integrity failure, got: {rep.failures}"
    )


def test_report_dataclass_defaults():
    """TableTestReport: passed=True, failures defaults to empty list."""
    rep = TableTestReport(passed=True)
    assert rep.failures == []
    rep2 = TableTestReport(passed=False, failures=["x"])
    assert not rep2.passed


def test_passes_single_row(tmp_path):
    """Boundary: single data row, sample covers it."""
    p = _wb(tmp_path, [["ID", "Score"], [1, 99]])
    h = split_workbook(p)[0]
    idx = build_index(p, h, row_key=["ID"])
    rep = run_table_tests(OpenpyxlFileSource(p), _canon(h), idx, sample_size=1)
    assert rep.passed


def test_row_integrity_detects_key_remap(tmp_path):
    """Row-integrity check must catch a row remap: _key_to_phys["Alice"] -> wrong row.

    If _key_to_phys["Alice"] points to row 3 (Bob's row), the index resolves "Alice"
    to the wrong physical row.  The row-integrity check reads the key-column cell at
    that physical row and finds "Bob" != "Alice", flagging a failure.
    """
    p = _wb(tmp_path, [["Name", "Val"], ["Alice", 10], ["Bob", 20]])
    h = split_workbook(p)[0]
    idx = build_index(p, h, row_key=["Name"])
    # Remap Alice to Bob's physical row
    idx._key_to_phys["Alice"] = idx._key_to_phys["Bob"]
    rep = run_table_tests(OpenpyxlFileSource(p), _canon(h), idx)
    assert not rep.passed, "Expected failure when key maps to wrong physical row"
    assert any("row-integrity" in f for f in rep.failures), (
        f"Expected row-integrity failure, got: {rep.failures}"
    )


def test_gate_flags_duplicate_columns(tmp_path):
    """
    Fix 3: fail-loud gate — duplicate column names in CanonicalTable.columns must
    produce a column-name failure even when the index itself is internally consistent.

    Scenario: table has columns [Region, Revenue] but we hand run_table_tests a
    CanonicalTable whose .columns list contains ["Region", "Revenue", "Revenue"]
    (duplicated — the corruption that col-axis fan-out used to produce).  The gate
    must flag this as a failure so it cannot escape with errors==[].
    """
    from mcg_swarm.schemas import CanonicalTable, ExtractionRef, ColumnSpec

    p = _wb(tmp_path, [["Region", "Revenue"], ["EMEA", 100], ["APAC", 200]])
    h = split_workbook(p)[0]
    idx = build_index(p, h, row_key=["Region"])

    # Build a CanonicalTable with a duplicated column name (simulates col-axis corruption)
    dup_columns = list(h.columns) + [h.columns[1]]  # Revenue duplicated
    table_with_dup = CanonicalTable(
        table_id="t",
        sheet=h.sheet,
        region=h.region,
        header_row=h.header_row,
        columns=dup_columns,
        description="d",
        extraction=ExtractionRef(script_name="idx", row_key=["Region"]),
    )

    rep = run_table_tests(OpenpyxlFileSource(p), table_with_dup, idx)

    assert not rep.passed, "Expected failure for duplicate column names"
    assert any("duplicate" in f.lower() or "column-name" in f.lower() for f in rep.failures), (
        f"Expected duplicate/column-name failure, got: {rep.failures}"
    )


def test_roundtrip_detects_query_divergence(tmp_path):
    """Phase 3 round-trip must catch when index.query() returns a different value than live_cache.

    We patch index.query() to return a wrong value for one cell, then verify run_table_tests
    flags a round-trip failure — proving Phase 3 is not a tautology.
    """
    from unittest.mock import patch, MagicMock

    p = _wb(tmp_path, [["Region", "Revenue"], ["EMEA", 100], ["APAC", 200]])
    h = split_workbook(p)[0]
    idx = build_index(p, h, row_key=["Region"])
    table = _canon(h)

    # Patch index.query to return a wrong value for "Revenue" column
    original_query = idx.query
    def bad_query(k, col):
        result = original_query(k, col)
        if col == "Revenue":
            bad = MagicMock()
            bad.value = 999999  # wrong value
            return bad
        return result

    with patch.object(idx, "query", side_effect=bad_query):
        rep = run_table_tests(OpenpyxlFileSource(p), table, idx)

    assert not rep.passed, f"Expected round-trip failure but passed: {rep.failures}"
    assert any("round-trip" in f for f in rep.failures), (
        f"Expected round-trip failure message, got: {rep.failures}"
    )


def test_gate_flags_column_name_mismatch_vs_live_header(tmp_path):
    """
    Fix 3: fail-loud gate — CanonicalTable.columns names that don't match the live
    header must be flagged.

    Scenario: workbook has [Region, Revenue] but table.columns claims [Region, Turnover].
    The gate must detect that "Turnover" is not present in the live header and flag it.
    """
    from mcg_swarm.schemas import CanonicalTable, ExtractionRef, ColumnSpec

    p = _wb(tmp_path, [["Region", "Revenue"], ["EMEA", 100], ["APAC", 200]])
    h = split_workbook(p)[0]
    idx = build_index(p, h, row_key=["Region"])

    # Build table claiming a column name that doesn't exist in the live header
    wrong_columns = [
        h.columns[0],  # Region — correct
        ColumnSpec(name="Turnover", dtype="number", role="value"),  # wrong name
    ]
    table_wrong_name = CanonicalTable(
        table_id="t",
        sheet=h.sheet,
        region=h.region,
        header_row=h.header_row,
        columns=wrong_columns,
        description="d",
        extraction=ExtractionRef(script_name="idx", row_key=["Region"]),
    )

    rep = run_table_tests(OpenpyxlFileSource(p), table_wrong_name, idx)

    assert not rep.passed, "Expected failure for column name mismatch vs live header"
    assert any(
        "column-name" in f.lower() or "not found" in f.lower() or "mismatch" in f.lower()
        for f in rep.failures
    ), f"Expected column-name/mismatch failure, got: {rep.failures}"


def test_gate_handles_region_with_empty_cells(tmp_path):
    """Regression: EmptyCell crash when region includes left-offset/sparse rows with empty cells.

    Reproduces the bug: openpyxl read_only mode returns EmptyCell objects for cells
    that were never written to the file.  EmptyCell has NO .row/.column attributes.
    The old batch-scan code did `pos = (cell.row, cell.column)` → AttributeError.

    Layout mimics capex_plan.xlsx: data starts at column B (col 2), so column A
    is entirely absent from the file.  The region "A2:C5" includes col A (scan_min_col=1)
    but ALL cells in column A are EmptyCell in read_only mode.

      row 2: [<empty>, "Key", "A", "B"]  ← header at col B-D but region includes A
      row 3: [<empty>, "x",   10,  20]
      row 4: [<empty>, "y",   30,  40]

    scan_min_col = 1 (col A) because that's the left edge of "A2:C5" wait — actually
    we model it as the table region being B2:D4 with col A absent, so the scan hits
    EmptyCell when iter_rows spans that absent column.

    The fix uses values_only=True + row/col offset arithmetic, avoiding cell objects.
    """
    import openpyxl as xl
    from mcg_swarm.schemas import CanonicalTable, ExtractionRef, ColumnSpec
    from mcg_swarm.extraction import ExtractionIndex
    from mcg_swarm.source import OpenpyxlFileSource

    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Data"
    # Simulate left-offset table: col A (col index 1) is never written.
    # Data lives in cols B-D (indices 2-4).
    # Row 1 is a banner row that only has content in col C (col index 3).
    # Only write cells that actually have data — leave col A and D1 absent.
    ws["C1"] = "Report Title"   # banner in C1 only; A1, B1, D1 absent → EmptyCell
    ws["B2"] = "Key"
    ws["C2"] = "Alpha"
    ws["D2"] = "Beta"
    ws["B3"] = "x"
    ws["C3"] = 10
    ws["D3"] = 20
    ws["B4"] = "y"
    ws["C4"] = 30
    ws["D4"] = 40

    path = str(tmp_path / "offset.xlsx")
    wb.save(path)

    # Verify EmptyCell is actually returned for A2 in read_only mode (pre-condition)
    wb_ro = xl.load_workbook(path, data_only=True, read_only=True)
    ws_ro = wb_ro["Data"]
    a2_row = list(ws_ro.iter_rows(min_row=2, max_row=2, min_col=1, max_col=1))[0]
    a2_cell = a2_row[0]
    assert type(a2_cell).__name__ == "EmptyCell", (
        f"Pre-condition failed: expected EmptyCell for A2, got {type(a2_cell).__name__}"
    )
    wb_ro.close()

    # Region A1:D4 covers the banner row AND the left-offset empty column A.
    # The scan bounding box will include col A → EmptyCell crash pre-fix.
    columns = [
        ColumnSpec(name="Key", dtype="string", role="key"),
        ColumnSpec(name="Alpha", dtype="number", role="value"),
        ColumnSpec(name="Beta", dtype="number", role="value"),
    ]
    table = CanonicalTable(
        table_id="offset_table",
        sheet="Data",
        region="A1:D4",
        header_row=2,
        columns=columns,
        description="offset table with empty left column in region",
        extraction=ExtractionRef(script_name="idx", row_key=["Key"]),
    )

    # Build index: ExtractionIndex reads via WorkbookSource (no EmptyCell bug there)
    idx = ExtractionIndex(OpenpyxlFileSource(path), "Data", "A1:D4", 2, columns, ["Key"])

    # This must NOT raise AttributeError on EmptyCell — should return a passing report
    rep = run_table_tests(OpenpyxlFileSource(path), table, idx)

    assert isinstance(rep, TableTestReport), f"Expected TableTestReport, got {type(rep)}"
    assert rep.passed, f"Expected gate to pass for correct table, failures: {rep.failures}"
