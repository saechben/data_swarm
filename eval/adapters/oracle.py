"""Reference oracle adapter.

It answers from the ground-truth labels, so it should score ~100% — proving the
labels are internally consistent and the scorer works. ``NoisyOracleAdapter``
injects controlled errors so you can see the scorer produce sub-100% breakdowns.

The oracle is *allowed* to read the labels; a real swarm adapter is not.
"""
from __future__ import annotations

import random
from typing import Any, Optional

from eval.adapters.base import DetectedMeasure, EvalAdapter, SemanticResult
from eval.schemas import WorkbookLabel
from eval.util import range_box, safe_eval
from openpyxl.utils import get_column_letter


class OracleAdapter(EvalAdapter):
    name = "oracle"

    def __init__(self) -> None:
        self._labels: dict[str, WorkbookLabel] = {}
        self._semantic: dict[str, dict[str, SemanticResult]] = {}
        self._cells: dict[str, dict[tuple[str, str, str], Any]] = {}

    def prepare(self, workbook_path: str, label: WorkbookLabel) -> None:
        wb = label.workbook
        self._labels[wb] = label
        self._cells[wb] = {
            (t.table_id, c.row_label, c.col_label): c.value
            for t in label.tables for c in t.cells
        }
        sem = {}
        for s in label.samples:
            if s.type == "semantic":
                sem[s.query] = SemanticResult(
                    value=s.expected_value, table_id=s.expected_table_id,
                    row_label=s.expected_row_label, col_label=s.expected_col_label,
                )
        self._semantic[wb] = sem

    # -- capability methods ---------------------------------------------------
    def table_region(self, wb, table_id, table_name, sheet) -> Optional[str]:
        try:
            return self._labels[wb].table(table_id).region
        except KeyError:
            return None

    def extract(self, wb, table_id, table_name, sheet, row_label, col_label) -> Any:
        return self._cells[wb].get((table_id, row_label, col_label))

    def answer_semantic(self, wb, query) -> SemanticResult:
        return self._semantic[wb].get(query, SemanticResult())

    def detected_measures(self, wb) -> list[DetectedMeasure]:
        out = []
        for m in self._labels[wb].measures:
            out.append(DetectedMeasure(
                table_id=m.table_id, row_label=m.row_label, col_label=m.col_label,
                value=m.value, semantic_name=m.semantic_name, aliases=list(m.aliases),
            ))
        return out

    def compute_formula(self, wb, expression, operands, business_logic) -> Optional[float]:
        label = self._labels[wb]
        names = {}
        for sym, mname in operands.items():
            names[sym] = float(label.measure(mname).value)
        return safe_eval(expression, names)


class NoisyOracleAdapter(OracleAdapter):
    """Oracle + controlled errors, to demonstrate the scorer discriminates."""

    name = "noisy"

    def __init__(self, error_rate: float = 0.18, seed: int = 7) -> None:
        super().__init__()
        self.error_rate = error_rate
        self._rng = random.Random(seed)

    def _flip(self) -> bool:
        return self._rng.random() < self.error_rate

    def table_region(self, wb, table_id, table_name, sheet):
        region = super().table_region(wb, table_id, table_name, sheet)
        if region and self._flip():
            # shrink the box by one row/col -> lowers IoU below threshold
            r1, c1, r2, c2 = range_box(region)
            r2 = max(r1, r2 - 1)
            return f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
        return region

    def extract(self, wb, table_id, table_name, sheet, row_label, col_label):
        v = super().extract(wb, table_id, table_name, sheet, row_label, col_label)
        if isinstance(v, (int, float)) and not isinstance(v, bool) and self._flip():
            return v + 1  # off-by-one extraction error
        return v

    def answer_semantic(self, wb, query):
        res = super().answer_semantic(wb, query)
        if isinstance(res.value, (int, float)) and self._flip():
            res = SemanticResult(value=res.value * 1.1, table_id=res.table_id,
                                 row_label=res.row_label, col_label=res.col_label)
        return res

    def detected_measures(self, wb):
        ms = super().detected_measures(wb)
        # occasionally drop a measure (recall hit) or corrupt a value (precision hit)
        out = []
        for m in ms:
            if self._flip():
                continue
            if self._flip():
                m.value = (m.value + 1) if isinstance(m.value, (int, float)) else m.value
            out.append(m)
        return out

    def compute_formula(self, wb, expression, operands, business_logic):
        v = super().compute_formula(wb, expression, operands, business_logic)
        if v is not None and self._flip():
            return v * 1.25
        return v
