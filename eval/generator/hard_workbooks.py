"""Extreme stress-test workbooks: the logic lives *inside* Excel.

These three files use live formulas, cross-sheet references, named ranges, chained
dependencies, SUMIF aggregation, two-level merged headers and display sign-formats.
A naive reader that ignores cached results or can't follow references will fail; a
correct extractor returns the recalculated value. Ground truth is computed here in
Python (clean numbers, exact), and the files are recalculated by LibreOffice in
build.py so they carry cached results like a real saved workbook.

Each builder returns (WorkbookLabel, needs_recalc=True).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from eval.generator.sampling import make_samples, prune_cells, resolve_measures
from eval.generator.specs import FormulaDef, MeasureDef
from eval.schemas import CellFact, ColumnLabel, RowKeyLabel, TableLabel, WorkbookLabel


class Grid:
    """Place a vertical table (optionally with a two-level grouped header).

    Addresses for every (row_label, col_label) are precomputed so formulas can
    reference any cell (including cross-sheet via ``ref(..., cross=True)``).
    """

    def __init__(self, ws, sheet, anchor, table_id, name, key_header, cols, rows,
                 groups=None):
        self.ws, self.sheet = ws, sheet
        self.table_id, self.name, self.key_header = table_id, name, key_header
        self.cols = cols          # list of (label, unit, dtype)
        self.rows = rows          # list of row labels
        self.groups = groups      # list aligned with cols (or None)
        self.r0, self.c0 = anchor
        self.header_rows = 2 if groups else 1
        self.data_top = self.r0 + self.header_rows
        self.addr = {}
        for j, (lbl, _, _) in enumerate(cols):
            for i, rl in enumerate(rows):
                self.addr[(rl, lbl)] = (
                    f"{get_column_letter(self.c0 + 1 + j)}{self.data_top + i}")
        self.cells: list[CellFact] = []

    def ref(self, row_label, col_label, cross=False):
        a = self.addr[(row_label, col_label)]
        return f"'{self.sheet}'!{a}" if cross else a

    def row_range(self, row_label, c_from, c_to, cross=False):
        a = f"{self.addr[(row_label, c_from)]}:{self.addr[(row_label, c_to)]}"
        return f"'{self.sheet}'!{a}" if cross else a

    def col_range(self, col_label, cross=False):
        top = self.addr[(self.rows[0], col_label)]
        bot = self.addr[(self.rows[-1], col_label)]
        return f"'{self.sheet}'!{top}:{bot}" if cross else f"{top}:{bot}"

    def write_header(self):
        ws = self.ws
        if self.groups:
            ws.cell(self.r0, self.c0, self.key_header)
            ws.merge_cells(start_row=self.r0, start_column=self.c0,
                           end_row=self.r0 + 1, end_column=self.c0)
            j = 0
            while j < len(self.cols):
                g = self.groups[j]
                k = j
                while k < len(self.cols) and self.groups[k] == g:
                    k += 1
                if g:
                    ws.cell(self.r0, self.c0 + 1 + j, g)
                    if k - j > 1:
                        ws.merge_cells(start_row=self.r0, start_column=self.c0 + 1 + j,
                                       end_row=self.r0, end_column=self.c0 + (k - j) + j)
                j = k
            for jj, (lbl, _, _) in enumerate(self.cols):
                ws.cell(self.r0 + 1, self.c0 + 1 + jj, lbl)
        else:
            ws.cell(self.r0, self.c0, self.key_header)
            for jj, (lbl, _, _) in enumerate(self.cols):
                ws.cell(self.r0, self.c0 + 1 + jj, lbl)

    def write(self, values, formulas=None, number_formats=None):
        """values: {(rl,cl): ground-truth}; formulas: {(rl,cl): excel text}."""
        formulas = formulas or {}
        nf = number_formats or {}
        ws = self.ws
        for i, rl in enumerate(self.rows):
            ws.cell(self.data_top + i, self.c0, rl)
            for j, (cl, _u, _d) in enumerate(self.cols):
                a = self.addr[(rl, cl)]
                val = values[(rl, cl)]
                if (rl, cl) in formulas:
                    ws[a] = formulas[(rl, cl)]
                    self.cells.append(CellFact(row_label=rl, col_label=cl, cell_ref=a,
                                               value=val, raw=formulas[(rl, cl)],
                                               is_formula=True))
                else:
                    ws[a] = val
                    self.cells.append(CellFact(row_label=rl, col_label=cl, cell_ref=a,
                                               value=val, raw=val))
                if (rl, cl) in nf:
                    ws[a].number_format = nf[(rl, cl)]

    def label(self, traps=None):
        gcl = get_column_letter
        last_col = self.c0 + len(self.cols)
        last_row = self.data_top + len(self.rows) - 1
        region = f"{gcl(self.c0)}{self.r0}:{gcl(last_col)}{last_row}"
        header_region = (f"{gcl(self.c0)}{self.r0}:{gcl(last_col)}"
                         f"{self.r0 + self.header_rows - 1}")
        data_region = f"{gcl(self.c0)}{self.data_top}:{gcl(last_col)}{last_row}"
        columns = [ColumnLabel(label=lbl, col_index=self.c0 + 1 + j,
                               col_letter=gcl(self.c0 + 1 + j), dtype=d, unit=u)
                   for j, (lbl, u, d) in enumerate(self.cols)]
        row_keys = [RowKeyLabel(label=rl, row_index=self.data_top + i)
                    for i, rl in enumerate(self.rows)]
        return TableLabel(table_id=self.table_id, name=self.name, sheet=self.sheet,
                          region=region, header_region=header_region,
                          data_region=data_region, columns=columns, row_keys=row_keys,
                          cells=self.cells, traps=traps or [])


def _finish(filename, difficulty, domain, traps, business_logic, tables, measure_defs,
            formula_defs):
    measures = resolve_measures(tables, measure_defs)
    samples, referenced = make_samples(filename, tables, measures, formula_defs,
                                       business_logic, prioritize_formulas=True)
    prune_cells(tables, referenced)
    sheets = list(dict.fromkeys(t.sheet for t in tables))
    return WorkbookLabel(
        workbook=filename, rel_path=f"workbooks/{filename}", difficulty=difficulty,
        domain=domain, traps=traps, sheets=sheets, business_logic=business_logic,
        tables=tables, measures=measures, samples=samples)


# --------------------------------------------------------------------------- #
# X1 — chained formulas + SUM ranges + a named range (TaxRate)
# --------------------------------------------------------------------------- #
def build_formula_chain(wb_dir: Path):
    fn = "formula_chain_pnl.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    qs = ["Q1", "Q2", "Q3", "Q4"]
    cols = [(q, "USD", "number") for q in qs] + [("FY", "USD", "number")]
    line_items = ["Units", "Price", "Revenue", "COGS", "Gross Profit", "OpEx",
                  "EBIT", "Tax", "Net Income"]
    g = Grid(ws, "Model", (1, 1), "pnl", "P&L Model", "Line Item", cols, line_items)

    units = {"Q1": 100, "Q2": 110, "Q3": 120, "Q4": 130}
    price = {q: 50 for q in qs}
    cogs = {"Q1": 2000, "Q2": 2200, "Q3": 2400, "Q4": 2600}
    opex = {"Q1": 1000, "Q2": 1300, "Q3": 1100, "Q4": 1400}
    tax_rate = 0.2

    revenue = {q: units[q] * price[q] for q in qs}
    gp = {q: revenue[q] - cogs[q] for q in qs}
    ebit = {q: gp[q] - opex[q] for q in qs}
    tax = {q: ebit[q] * tax_rate for q in qs}
    ni = {q: ebit[q] - tax[q] for q in qs}
    gt = {"Units": units, "Price": price, "Revenue": revenue, "COGS": cogs,
          "Gross Profit": gp, "OpEx": opex, "EBIT": ebit, "Tax": tax, "Net Income": ni}

    # TaxRate named range -> a cell on the sheet
    ws["H1"] = "TaxRate"
    ws["I1"] = tax_rate
    wb.defined_names.add(DefinedName("TaxRate", attr_text="Model!$I$1"))

    values, formulas = {}, {}
    for li in line_items:
        for q in qs:
            values[(li, q)] = gt[li][q]
        values[(li, "FY")] = sum(gt[li][q] for q in qs)
    for q in qs:
        formulas[("Revenue", q)] = f"={g.ref('Units', q)}*{g.ref('Price', q)}"
        formulas[("Gross Profit", q)] = f"={g.ref('Revenue', q)}-{g.ref('COGS', q)}"
        formulas[("EBIT", q)] = f"={g.ref('Gross Profit', q)}-{g.ref('OpEx', q)}"
        formulas[("Tax", q)] = f"={g.ref('EBIT', q)}*TaxRate"
        formulas[("Net Income", q)] = f"={g.ref('EBIT', q)}-{g.ref('Tax', q)}"
    for li in line_items:
        formulas[(li, "FY")] = f"=SUM({g.row_range(li, 'Q1', 'Q4')})"

    g.write_header()
    g.write(values, formulas)
    table = g.label(traps=["chained_formulas", "sum_range", "named_range",
                           "totals_column_formula"])
    wb.save(wb_dir / fn)

    measure_defs = [
        MeasureDef("revenue_q1", "pnl", "Revenue", "Q1"),
        MeasureDef("gross_profit_q1", "pnl", "Gross Profit", "Q1"),
        MeasureDef("opex_q1", "pnl", "OpEx", "Q1"),
        MeasureDef("ebit_q1", "pnl", "EBIT", "Q1"),
        MeasureDef("net_income_q1", "pnl", "Net Income", "Q1"),
        MeasureDef("net_income_fy", "pnl", "Net Income", "FY"),
    ]
    formula_defs = [
        FormulaDef("EBIT Q1 from gross profit and opex", "G - O",
                   {"G": "gross_profit_q1", "O": "opex_q1"}),
        FormulaDef("Net income margin Q1 (%)", "N / R * 100",
                   {"N": "net_income_q1", "R": "revenue_q1"}),
    ]
    bl = ("Revenue = units x price; Gross profit = revenue - COGS; "
          "EBIT = gross profit - OpEx; Tax = EBIT x TaxRate; "
          "Net income = EBIT - Tax. FY = sum of quarters.")
    label = _finish(fn, "hard", "finance",
                    ["chained_formulas", "sum_range", "named_range"], bl,
                    [table], measure_defs, formula_defs)
    return label, True


# --------------------------------------------------------------------------- #
# X2 — cross-sheet references + named range (FXRate) across 3 sheets
# --------------------------------------------------------------------------- #
def build_cross_sheet(wb_dir: Path):
    fn = "cross_sheet_model.xlsx"
    wb = Workbook()
    inp = wb.active
    inp.title = "Inputs"
    calc = wb.create_sheet("Calc")
    summ = wb.create_sheet("Summary")
    regions = ["EMEA", "APAC", "NorthAm"]

    units = {"EMEA": 200, "APAC": 150, "NorthAm": 300}
    asp = {"EMEA": 100, "APAC": 120, "NorthAm": 90}
    cost = {"EMEA": 60, "APAC": 90, "NorthAm": 54}
    fx = 1.1

    gi = Grid(inp, "Inputs", (1, 1), "inputs", "Regional Drivers", "Region",
              [("Units", "count", "number"), ("ASP", "USD", "number"),
               ("UnitCost", "USD", "number")], regions)
    gi.write_header()
    gi.write({(r, "Units"): units[r] for r in regions} |
             {(r, "ASP"): asp[r] for r in regions} |
             {(r, "UnitCost"): cost[r] for r in regions})
    inp["F1"] = "FXRate"
    inp["G1"] = fx
    wb.defined_names.add(DefinedName("FXRate", attr_text="Inputs!$G$1"))

    rev_local = {r: units[r] * asp[r] for r in regions}
    margin = {r: (asp[r] - cost[r]) / asp[r] for r in regions}
    rev_usd = {r: rev_local[r] * fx for r in regions}

    gc = Grid(calc, "Calc", (1, 1), "calc", "Regional Calc", "Region",
              [("RevenueLocal", "local", "number"), ("Margin", "ratio", "number"),
               ("RevenueUSD", "USD", "number")], regions)
    cvals, cform = {}, {}
    for r in regions:
        cvals[(r, "RevenueLocal")] = rev_local[r]
        cvals[(r, "Margin")] = margin[r]
        cvals[(r, "RevenueUSD")] = rev_usd[r]
        cform[(r, "RevenueLocal")] = f"={gi.ref(r, 'Units', True)}*{gi.ref(r, 'ASP', True)}"
        cform[(r, "Margin")] = (f"=({gi.ref(r, 'ASP', True)}-{gi.ref(r, 'UnitCost', True)})"
                                f"/{gi.ref(r, 'ASP', True)}")
        cform[(r, "RevenueUSD")] = f"={gc.ref(r, 'RevenueLocal')}*FXRate"
    gc.write_header()
    gc.write(cvals, cform)

    gs = Grid(summ, "Summary", (1, 1), "summary", "Company Summary", "Entity",
              [("TotalRevenueUSD", "USD", "number")], ["Company"])
    total_usd = sum(rev_usd.values())
    gs.write_header()
    gs.write({("Company", "TotalRevenueUSD"): total_usd},
             {("Company", "TotalRevenueUSD"):
              f"=SUM({gc.col_range('RevenueUSD', True)})"})

    tables = [gi.label(traps=["named_range"]),
              gc.label(traps=["cross_sheet_refs", "named_range", "ratio_formula"]),
              gs.label(traps=["cross_sheet_refs", "sum_range"])]
    wb.save(wb_dir / fn)

    measure_defs = [
        MeasureDef("revenue_local_emea", "calc", "EMEA", "RevenueLocal"),
        MeasureDef("revenue_usd_emea", "calc", "EMEA", "RevenueUSD"),
        MeasureDef("revenue_usd_apac", "calc", "APAC", "RevenueUSD"),
        MeasureDef("revenue_usd_northam", "calc", "NorthAm", "RevenueUSD"),
        MeasureDef("total_revenue_usd", "summary", "Company", "TotalRevenueUSD"),
        MeasureDef("asp_emea", "inputs", "EMEA", "ASP"),
        MeasureDef("cost_emea", "inputs", "EMEA", "UnitCost"),
    ]
    formula_defs = [
        FormulaDef("Company revenue USD = sum of regional USD revenue", "E + A + N",
                   {"E": "revenue_usd_emea", "A": "revenue_usd_apac",
                    "N": "revenue_usd_northam"}),
        FormulaDef("EMEA margin (%)", "(A - C) / A * 100",
                   {"A": "asp_emea", "C": "cost_emea"}),
    ]
    bl = ("Revenue (local) = units x ASP; margin = (ASP - unit cost) / ASP; "
          "revenue (USD) = revenue (local) x FX rate; company total = sum across "
          "regions. Drivers live on Inputs; Calc and Summary reference them.")
    label = _finish(fn, "hard", "finance",
                    ["cross_sheet_refs", "named_range", "multi_sheet", "ratio_formula"],
                    bl, tables, measure_defs, formula_defs)
    return label, True


# --------------------------------------------------------------------------- #
# X3 — the kitchen sink: two-level header + cross-sheet formulas + SUMIF +
#      display sign-format + percentage formula + offset anchor + footnote
# --------------------------------------------------------------------------- #
def build_messy_everything(wb_dir: Path):
    fn = "messy_everything.xlsx"
    wb = Workbook()
    data = wb.active
    data.title = "Data"
    dash = wb.create_sheet("Dashboard")
    products = ["Product A", "Product B", "Product C"]

    units = {"Product A": 100, "Product B": 200, "Product C": 150}
    price = {"Product A": 10, "Product B": 5, "Product C": 8}
    returns = {"Product A": -150, "Product B": -250, "Product C": 0}

    gi = Grid(data, "Data", (1, 1), "inputs", "Product Inputs", "Product",
              [("Units", "count", "number"), ("Price", "USD", "number"),
               ("ReturnsValue", "USD", "number")], products)
    gi.write_header()
    gi.write(
        {(p, "Units"): units[p] for p in products} |
        {(p, "Price"): price[p] for p in products} |
        {(p, "ReturnsValue"): returns[p] for p in products},
        number_formats={(p, "ReturnsValue"): "#,##0;(#,##0)" for p in products},
    )

    # small ledger for SUMIF
    ledger_rows = [
        ("Product A", 120), ("Product B", 80), ("Product C", 200), ("Product A", 30),
        ("Product B", 70), ("Product C", 0), ("Product A", 50), ("Product C", 40),
        ("Product B", 60), ("Product A", 25),
    ]
    led_top = 7
    data.cell(led_top, 1, "Product")
    data.cell(led_top, 2, "Amount")
    for i, (p, amt) in enumerate(ledger_rows):
        data.cell(led_top + 1 + i, 1, p)
        data.cell(led_top + 1 + i, 2, amt)
    led_prod_range = f"Data!$A${led_top+1}:$A${led_top+len(ledger_rows)}"
    led_amt_range = f"Data!$B${led_top+1}:$B${led_top+len(ledger_rows)}"
    ledger_total = {p: sum(a for q, a in ledger_rows if q == p) for p in products}

    # Dashboard: offset anchor, two-level header, cross-sheet formulas, %, SUMIF
    gross = {p: units[p] * price[p] for p in products}
    net = {p: gross[p] + returns[p] for p in products}
    margin_pct = {p: net[p] / gross[p] * 100 for p in products}

    gd = Grid(dash, "Dashboard", (3, 2), "metrics", "Product Metrics", "Product",
              [("Gross", "USD", "number"), ("Net", "USD", "number"),
               ("MarginPct", "%", "number"), ("LedgerTotal", "USD", "number")],
              products,
              groups=["Revenue", "Revenue", None, None])
    dvals, dform = {}, {}
    for i, p in enumerate(products):
        dvals[(p, "Gross")] = gross[p]
        dvals[(p, "Net")] = net[p]
        dvals[(p, "MarginPct")] = margin_pct[p]
        dvals[(p, "LedgerTotal")] = ledger_total[p]
        # the product label cell in the key column (used as the SUMIF criterion)
        key_cell = f"{get_column_letter(gd.c0)}{gd.data_top + i}"
        dform[(p, "Gross")] = f"={gi.ref(p, 'Units', True)}*{gi.ref(p, 'Price', True)}"
        dform[(p, "Net")] = f"={gd.ref(p, 'Gross')}+{gi.ref(p, 'ReturnsValue', True)}"
        dform[(p, "MarginPct")] = f"={gd.ref(p, 'Net')}/{gd.ref(p, 'Gross')}*100"
        dform[(p, "LedgerTotal")] = (
            f"=SUMIF({led_prod_range},{key_cell},{led_amt_range})")
    gd.write_header()
    gd.write(dvals, dform,
             number_formats={(p, "Net"): "#,##0;(#,##0)" for p in products})

    # footnote marker + footnote table on Data
    data.cell(20, 1, "Note (a): Returns recorded as negative.")

    tables = [
        gi.label(traps=["display_paren_negatives"]),
        gd.label(traps=["two_level_header", "merged_cells", "cross_sheet_refs",
                        "sumif", "percentage_formula", "offset_anchor",
                        "display_paren_negatives"]),
    ]
    wb.save(wb_dir / fn)

    measure_defs = [
        MeasureDef("gross_a", "metrics", "Product A", "Gross"),
        MeasureDef("net_a", "metrics", "Product A", "Net"),
        MeasureDef("marginpct_a", "metrics", "Product A", "MarginPct"),
        MeasureDef("ledgertotal_a", "metrics", "Product A", "LedgerTotal"),
        MeasureDef("net_b", "metrics", "Product B", "Net"),
        MeasureDef("net_c", "metrics", "Product C", "Net"),
    ]
    formula_defs = [
        FormulaDef("Product A net margin (%)", "N / G * 100",
                   {"N": "net_a", "G": "gross_a"}),
        FormulaDef("Total net revenue across products", "A + B + C",
                   {"A": "net_a", "B": "net_b", "C": "net_c"}),
    ]
    bl = ("Gross = units x price; Net = gross + returns (returns are negative); "
          "margin % = net / gross x 100; ledger total = SUMIF of the transaction "
          "ledger by product. Dashboard cells reference the Data sheet.")
    label = _finish(fn, "hard", "ops",
                    ["two_level_header", "cross_sheet_refs", "sumif",
                     "percentage_formula", "display_paren_negatives", "offset_anchor"],
                    bl, tables, measure_defs, formula_defs)
    return label, True


BUILDERS = [build_formula_chain, build_cross_sheet, build_messy_everything]


def build_all(wb_dir: Path):
    return [b(wb_dir) for b in BUILDERS]
