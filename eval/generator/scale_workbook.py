"""Extreme-scale workbook: a fact table too big for a single pass.

``enterprise_transactions.xlsx`` is a ~100,000-row x 22-column transaction fact
table on one sheet, plus three derived summary tables (by region / month / category)
on their own sheets. At ~2.2M cells the primary sheet cannot be analysed in a single
pass — the orchestrator must segment it into row bands, so **at least two
subagents** are required just for that one table (and more across the summary sheets).

Written with openpyxl ``write_only`` (streaming, low memory). Ground truth: the
summary aggregates are accumulated while streaming, and re-checked from the file by
``verify.py``'s streaming pass. The label only materialises a bounded sample of cells
and row keys (the giant table is never fully serialised).
"""
from __future__ import annotations

import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from eval.generator.sampling import make_samples, resolve_measures
from eval.generator.specs import FormulaDef, MeasureDef
from eval.schemas import (
    CellFact, ColumnLabel, ExtractionSample, RowKeyLabel, TableLabel, WorkbookLabel,
)

N_ROWS = 100_000

REGIONS = ["EMEA", "APAC", "NorthAm", "LatAm", "MEA"]
COUNTRIES = {
    "EMEA": ["UK", "Germany", "France", "Spain"],
    "APAC": ["Japan", "Australia", "India", "Singapore"],
    "NorthAm": ["USA", "Canada", "Mexico"],
    "LatAm": ["Brazil", "Argentina", "Chile"],
    "MEA": ["UAE", "SouthAfrica", "Egypt"],
}
CHANNELS = ["Online", "Retail", "Partner", "Direct"]
SEGMENTS = ["Enterprise", "SMB", "Consumer"]
CATEGORIES = ["Widgets", "Gadgets", "Gizmos", "Sprockets", "Cogs", "Modules"]
CAT_RATIO = {"Widgets": 0.55, "Gadgets": 0.62, "Gizmos": 0.48, "Sprockets": 0.70,
             "Cogs": 0.40, "Modules": 0.58}
CAT_PRICE = {"Widgets": (20, 60), "Gadgets": (40, 120), "Gizmos": (60, 200),
             "Sprockets": (5, 25), "Cogs": (3, 15), "Modules": (80, 300)}
REPS = [f"Rep{n:02d}" for n in range(1, 21)]
MONTHS = [f"2024-{m:02d}" for m in range(1, 13)]

# A1 = TxnID (key). Value columns B..V:
COLUMNS = [
    ("Date", None, "date"), ("FiscalQuarter", None, "string"),
    ("Region", None, "string"), ("Country", None, "string"),
    ("Channel", None, "string"), ("Segment", None, "string"),
    ("ProductCategory", None, "string"), ("ProductSKU", None, "string"),
    ("SalesRep", None, "string"), ("CustomerID", None, "string"),
    ("Currency", None, "string"), ("UnitsSold", "count", "number"),
    ("GrossAmount", "USD", "number"), ("DiscountPct", "%", "number"),
    ("DiscountAmount", "USD", "number"), ("NetAmount", "USD", "number"),
    ("COGS", "USD", "number"), ("GrossProfit", "USD", "number"),
    ("TaxAmount", "USD", "number"), ("NetRevenue", "USD", "number"),
    ("Status", None, "string"),
]
COL_INDEX = {name: i + 2 for i, (name, _u, _d) in enumerate(COLUMNS)}  # B=2 ...
NEEDLE_NUMERIC = ["UnitsSold", "GrossAmount", "NetAmount", "GrossProfit", "NetRevenue"]


def _zero_agg():
    return {"UnitsSold": 0, "GrossAmount": 0.0, "NetAmount": 0.0, "COGS": 0.0,
            "GrossProfit": 0.0, "TaxAmount": 0.0, "NetRevenue": 0.0, "count": 0}


def build(wb_dir: Path):
    fn = "enterprise_transactions.xlsx"
    rng = random.Random(20240618)

    # rows (2-based; row 1 is header) chosen as "needles" to spot-check deep in the file
    needle_rows = sorted({rng.randint(2, N_ROWS + 1) for _ in range(45)} |
                         {2, N_ROWS + 1, N_ROWS // 2 + 1, 73422, 99987})
    needle_set = set(needle_rows)
    needle_cells: dict[int, dict] = {}

    reg_agg = {r: _zero_agg() for r in REGIONS}
    mon_agg = {m: _zero_agg() for m in MONTHS}
    cat_agg = {c: _zero_agg() for c in CATEGORIES}

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Transactions")
    ws.append(["TxnID"] + [name for name, _u, _d in COLUMNS])

    for i in range(N_ROWS):
        sheet_row = i + 2
        region = REGIONS[rng.randrange(len(REGIONS))]
        country = COUNTRIES[region][rng.randrange(len(COUNTRIES[region]))]
        cat = CATEGORIES[rng.randrange(len(CATEGORIES))]
        month = MONTHS[rng.randrange(12)]
        q = f"Q{(int(month[5:7]) - 1) // 3 + 1}"
        lo, hi = CAT_PRICE[cat]
        price = rng.randint(lo, hi)
        units = rng.randint(1, 50)
        gross = units * price
        disc_pct = rng.choice([0, 5, 10, 15, 20])
        disc_amt = round(gross * disc_pct / 100, 2)
        net = round(gross - disc_amt, 2)
        cogs = round(net * CAT_RATIO[cat], 2)
        gp = round(net - cogs, 2)
        tax = round(max(gp, 0) * 0.2, 2)
        netrev = net
        txn_id = f"T{i:06d}"
        sku = f"{cat[:3].upper()}-{rng.randint(1000, 9999)}"
        rep = REPS[rng.randrange(len(REPS))]
        cust = f"C{rng.randint(10000, 99999)}"
        status = rng.choice(["Closed", "Closed", "Closed", "Pending", "Refunded"])

        ws.append([txn_id, month + "-15", q, region, country,
                   rng.choice(CHANNELS), rng.choice(SEGMENTS), cat, sku, rep, cust,
                   "USD", units, gross, disc_pct, disc_amt, net, cogs, gp, tax,
                   netrev, status])

        for agg in (reg_agg[region], mon_agg[month], cat_agg[cat]):
            agg["UnitsSold"] += units
            agg["GrossAmount"] += gross
            agg["NetAmount"] += net
            agg["COGS"] += cogs
            agg["GrossProfit"] += gp
            agg["TaxAmount"] += tax
            agg["NetRevenue"] += netrev
            agg["count"] += 1

        if sheet_row in needle_set:
            needle_cells[sheet_row] = {
                "TxnID": txn_id, "UnitsSold": units, "GrossAmount": gross,
                "NetAmount": net, "GrossProfit": gp, "NetRevenue": netrev}

    # ---- summary sheets (small; appended after the giant sheet) --------------
    summaries = []
    summaries.append(_write_summary(
        wb, "Region Summary", "region_summary", "Regional Summary", "Region",
        [("UnitsSold", "count"), ("NetRevenue", "USD"), ("COGS", "USD"),
         ("GrossProfit", "USD"), ("TaxAmount", "USD")],
        REGIONS, reg_agg))
    summaries.append(_write_summary(
        wb, "Monthly Summary", "monthly_summary", "Monthly Summary", "Month",
        [("UnitsSold", "count"), ("NetRevenue", "USD"), ("GrossProfit", "USD")],
        MONTHS, mon_agg))
    summaries.append(_write_summary(
        wb, "Category Summary", "category_summary", "Category Summary", "Category",
        [("UnitsSold", "count"), ("NetRevenue", "USD"), ("COGS", "USD"),
         ("GrossProfit", "USD")],
        CATEGORIES, cat_agg))

    wb_dir.mkdir(parents=True, exist_ok=True)
    wb.save(wb_dir / fn)

    # ---- giant TableLabel (bounded cells + row keys) ------------------------
    last_col = get_column_letter(1 + len(COLUMNS))  # V
    last_row = N_ROWS + 1
    columns = [ColumnLabel(label=name, col_index=COL_INDEX[name],
                           col_letter=get_column_letter(COL_INDEX[name]),
                           dtype=d, unit=u) for name, u, d in COLUMNS]
    cells, row_keys = [], []
    for sr in needle_rows:
        nc = needle_cells[sr]
        row_keys.append(RowKeyLabel(label=nc["TxnID"], row_index=sr))
        for col in NEEDLE_NUMERIC:
            cells.append(CellFact(
                row_label=nc["TxnID"], col_label=col,
                cell_ref=f"{get_column_letter(COL_INDEX[col])}{sr}",
                value=nc[col], raw=nc[col]))
    giant = TableLabel(
        table_id="transactions", name="Transaction Ledger", sheet="Transactions",
        region=f"A1:{last_col}{last_row}",
        header_region=f"A1:{last_col}1",
        data_region=f"A2:{last_col}{last_row}",
        columns=columns, row_keys=row_keys, cells=cells,
        traps=["extreme_scale", "requires_segmentation", "100k_rows", "wide_table"])

    tables = [giant] + summaries

    measure_defs = []
    for r in REGIONS:
        measure_defs.append(MeasureDef(f"netrev_{r.lower()}", "region_summary", r,
                                       "NetRevenue", [f"{r} net revenue"]))
    measure_defs += [
        MeasureDef("gp_emea", "region_summary", "EMEA", "GrossProfit"),
        MeasureDef("gp_apac", "region_summary", "APAC", "GrossProfit"),
        MeasureDef("netrev_widgets", "category_summary", "Widgets", "NetRevenue"),
        MeasureDef("netrev_jan", "monthly_summary", "2024-01", "NetRevenue"),
    ]
    formula_defs = [
        FormulaDef("Company-wide NetRevenue = sum of regional NetRevenue",
                   "E + A + N + L + M",
                   {"E": "netrev_emea", "A": "netrev_apac", "N": "netrev_northam",
                    "L": "netrev_latam", "M": "netrev_mea"}),
        FormulaDef("EMEA gross margin (%)", "G / R * 100",
                   {"G": "gp_emea", "R": "netrev_emea"}),
    ]
    business_logic = (
        "Per transaction: gross = units x price; net = gross - discount; "
        "COGS = net x category ratio; gross profit = net - COGS; tax = 20% of "
        "positive gross profit; net revenue = net. Regional / monthly / category "
        "summaries aggregate the ledger; company net revenue = sum of regions.")

    measures = resolve_measures(tables, measure_defs)
    samples, _ref = make_samples(fn, tables, measures, formula_defs, business_logic)

    # guarantee a few deep-needle extraction samples (the haystack point)
    deep = sorted(needle_rows)[-3:]
    for k, sr in enumerate(deep):
        nc = needle_cells[sr]
        samples.append(ExtractionSample(
            id=f"{fn}:deep:{k}", sheet="Transactions", table_id="transactions",
            table="Transaction Ledger", row_label=nc["TxnID"], col_label="NetRevenue",
            expected_value=nc["NetRevenue"],
            expected_cell_ref=f"{get_column_letter(COL_INDEX['NetRevenue'])}{sr}",
            dtype="number"))

    label = WorkbookLabel(
        workbook=fn, rel_path=f"workbooks/{fn}", difficulty="hard", domain="sales",
        traps=["extreme_scale", "requires_segmentation", "100k_rows", "multi_sheet",
               "wide_table"],
        sheets=["Transactions", "Region Summary", "Monthly Summary",
                "Category Summary"],
        business_logic=business_logic, tables=tables, measures=measures,
        samples=samples)
    return label, False  # no in-cell formulas -> no LibreOffice recalc


def _write_summary(wb, sheet, table_id, name, key_header, col_specs, row_labels, agg):
    """Append a small summary table (write_only) and build its TableLabel."""
    ws = wb.create_sheet(sheet)
    ws.append([key_header] + [c for c, _u in col_specs])
    cells, row_keys = [], []
    for i, rl in enumerate(row_labels):
        sheet_row = i + 2
        ws.append([rl] + [agg[rl][c] for c, _u in col_specs])
        row_keys.append(RowKeyLabel(label=rl, row_index=sheet_row))
        for j, (c, _u) in enumerate(col_specs):
            cells.append(CellFact(row_label=rl, col_label=c,
                                  cell_ref=f"{get_column_letter(2 + j)}{sheet_row}",
                                  value=agg[rl][c], raw=agg[rl][c]))
    last_col = get_column_letter(1 + len(col_specs))
    last_row = 1 + len(row_labels)
    columns = [ColumnLabel(label=c, col_index=2 + j, col_letter=get_column_letter(2 + j),
                           dtype="number", unit=u) for j, (c, u) in enumerate(col_specs)]
    return TableLabel(
        table_id=table_id, name=name, sheet=sheet, region=f"A1:{last_col}{last_row}",
        header_region=f"A1:{last_col}1", data_region=f"A2:{last_col}{last_row}",
        columns=columns, row_keys=row_keys, cells=cells, traps=["aggregate_table"])
