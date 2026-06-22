"""In-memory table model + renderer.

A ``TableSpec`` is the single source of truth for one table: the generator writes
it into an xlsx worksheet AND derives the ``TableLabel`` (regions, columns, row
keys, per-cell facts) from the exact same object. That is what makes the ground
truth correct by construction.

Supported realistic "mess":
  - title banner merged across the table width
  - offset anchor (table not at A1)
  - single or two-level (grouped, merged) column headers
  - units in the header text or in a dedicated units row
  - totals row and/or totals column
  - footnote markers in cells   ("1234 (a)"   -> value 1234)
  - thousands-formatted text     ("1,234"      -> value 1234)
  - parenthesised negatives      ("(123)"      -> value -123)
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Optional

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from eval.schemas import CellFact, ColumnLabel, RowKeyLabel, TableLabel


@dataclass
class ColSpec:
    label: str
    unit: Optional[str] = None
    dtype: str = "number"
    group: Optional[str] = None  # set on >=1 col to trigger a two-level header


@dataclass
class TableSpec:
    table_id: str
    name: str
    sheet: str
    anchor: tuple[int, int]  # (row, col), 1-based, top-left of the written block
    key_header: str  # header of the row-key (left) column
    columns: list[ColSpec]
    rows: list[str]  # row-key labels
    data: dict[tuple[str, str], object]  # (row_label, col_label) -> ground-truth value

    title: Optional[str] = None  # merged banner above the header
    units_in_header: bool = False
    units_row: bool = False
    totals_row: Optional[str] = None  # label, e.g. "Total"
    totals_col: Optional[str] = None  # label, e.g. "Total"

    footnote_cells: set[tuple[str, str]] = field(default_factory=set)
    thousands_text: bool = False  # write numerics as comma text
    paren_negatives: bool = False  # write negatives as "(n)" text

    orientation: str = "vertical"
    traps: list[str] = field(default_factory=list)
    is_duplicate_of: Optional[str] = None


def _fmt_raw(value, *, footnote: bool, thousands: bool, paren_neg: bool):
    """Return what is literally written to the cell for a numeric value."""
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        return value
    if paren_neg and value < 0:
        body = f"{abs(value):,}" if thousands else f"{abs(value):g}"
        return f"({body})"
    if thousands:
        txt = f"{value:,}"
        return f"{txt} (a)" if footnote else txt
    if footnote:
        return f"{value:g} (a)"
    return value


def render_table(ws: Worksheet, spec: TableSpec) -> TableLabel:
    """Write ``spec`` into ``ws`` and return its ground-truth ``TableLabel``."""
    r0, c0 = spec.anchor
    leaf_cols = list(spec.columns)
    if spec.totals_col:
        leaf_cols = leaf_cols + [ColSpec(spec.totals_col, dtype="number")]

    n_value_cols = len(leaf_cols)
    width = 1 + n_value_cols  # key col + value cols
    last_col = c0 + width - 1

    grouped = any(c.group for c in leaf_cols)
    row = r0

    # --- title banner ------------------------------------------------------
    if spec.title:
        ws.cell(row=row, column=c0, value=spec.title)
        ws.merge_cells(
            start_row=row, start_column=c0, end_row=row, end_column=last_col
        )
        row += 1

    # --- header(s) ---------------------------------------------------------
    header_top = row
    if grouped:
        # top row: group labels merged across consecutive same-group columns
        ws.cell(row=row, column=c0, value=spec.key_header)
        ws.merge_cells(start_row=row, start_column=c0, end_row=row + 1, end_column=c0)
        cc = c0 + 1
        i = 0
        while i < len(leaf_cols):
            grp = leaf_cols[i].group
            j = i
            while j < len(leaf_cols) and leaf_cols[j].group == grp:
                j += 1
            span = j - i
            if grp:
                ws.cell(row=row, column=cc, value=grp)
                if span > 1:
                    ws.merge_cells(
                        start_row=row, start_column=cc,
                        end_row=row, end_column=cc + span - 1,
                    )
            cc += span
            i = j
        row += 1
        # sub header
        for k, col in enumerate(leaf_cols):
            label = col.label
            if spec.units_in_header and col.unit:
                label = f"{label} ({col.unit})"
            ws.cell(row=row, column=c0 + 1 + k, value=label)
        header_bottom = row
        row += 1
    else:
        ws.cell(row=row, column=c0, value=spec.key_header)
        for k, col in enumerate(leaf_cols):
            label = col.label
            if spec.units_in_header and col.unit:
                label = f"{label} ({col.unit})"
            ws.cell(row=row, column=c0 + 1 + k, value=label)
        header_bottom = row
        row += 1

    # --- optional units row ------------------------------------------------
    if spec.units_row:
        ws.cell(row=row, column=c0, value="")
        for k, col in enumerate(leaf_cols):
            ws.cell(row=row, column=c0 + 1 + k, value=(col.unit or ""))
        row += 1

    # --- data rows ---------------------------------------------------------
    data_top = row
    cells: list[CellFact] = []
    row_keys: list[RowKeyLabel] = []

    body_rows = list(spec.rows)
    if spec.totals_row:
        body_rows = body_rows + [spec.totals_row]

    for rlabel in body_rows:
        ws.cell(row=row, column=c0, value=rlabel)
        row_keys.append(RowKeyLabel(label=rlabel, row_index=row))
        for k, col in enumerate(leaf_cols):
            clabel = col.label
            col_idx = c0 + 1 + k
            value = _cell_value(spec, rlabel, clabel, leaf_cols)
            if value is None:
                continue  # blank cell (e.g. total of a text column); no ground truth
            footnote = (rlabel, clabel) in spec.footnote_cells
            raw = _fmt_raw(
                value,
                footnote=footnote,
                thousands=spec.thousands_text,
                paren_neg=spec.paren_negatives,
            )
            ws.cell(row=row, column=col_idx, value=raw)
            cells.append(
                CellFact(
                    row_label=rlabel,
                    col_label=clabel,
                    cell_ref=f"{get_column_letter(col_idx)}{row}",
                    value=value,
                    raw=raw,
                )
            )
        row += 1

    data_bottom = row - 1
    last_row = data_bottom

    # --- regions -----------------------------------------------------------
    def rng(r1, c1, r2, c2):
        return f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"

    region = rng(r0, c0, last_row, last_col)
    header_region = rng(header_top, c0, header_bottom, last_col)
    data_region = rng(data_top, c0, data_bottom, last_col)

    columns = [
        ColumnLabel(
            label=col.label,
            col_index=c0 + 1 + k,
            col_letter=get_column_letter(c0 + 1 + k),
            dtype=col.dtype,  # type: ignore[arg-type]
            unit=col.unit,
        )
        for k, col in enumerate(leaf_cols)
    ]

    return TableLabel(
        table_id=spec.table_id,
        name=spec.name,
        sheet=spec.sheet,
        region=region,
        header_region=header_region,
        data_region=data_region,
        orientation=spec.orientation,  # type: ignore[arg-type]
        columns=columns,
        row_keys=row_keys,
        cells=cells,
        traps=list(spec.traps),
        is_duplicate_of=spec.is_duplicate_of,
    )


def _cell_value(spec: TableSpec, rlabel: str, clabel: str, leaf_cols: list[ColSpec]):
    """Resolve a cell's ground-truth value, computing totals where needed."""
    is_total_row = bool(spec.totals_row) and rlabel == spec.totals_row
    is_total_col = bool(spec.totals_col) and clabel == spec.totals_col

    if is_total_row and is_total_col:
        # grand total: sum over real rows x real numeric columns
        return sum(
            _base_value(spec, rr, c.label)
            for rr in spec.rows
            for c in spec.columns
            if c.dtype == "number"
        )
    if is_total_col:
        # row sum across the real (non-total) numeric columns
        return sum(
            _base_value(spec, rlabel, c.label)
            for c in spec.columns
            if c.dtype == "number"
        )
    if is_total_row:
        # column sum across the real (non-total) rows
        col = next(c for c in leaf_cols if c.label == clabel)
        if col.dtype != "number":
            return None  # no total for a text column -> cell left blank, no fact
        return sum(_base_value(spec, rr, clabel) for rr in spec.rows)
    return _base_value(spec, rlabel, clabel)


def _base_value(spec: TableSpec, rlabel: str, clabel: str):
    return spec.data[(rlabel, clabel)]
