#!/usr/bin/env python3
"""Verify that every label is correct against the physical xlsx.

This is what guarantees "the expected result has to be correct". For each workbook:
  - every CellFact's literal text matches the cell in the file
  - every CellFact's clean value is recoverable from that literal text
  - every region lies within the sheet and matches its cells
  - every measure resolves to its labeled cell value
  - every sample's expected value matches the file (extraction/semantic), the
    table region (boundary), or recomputes from operands (formula)
  - the large-ledger aggregates are recomputed from all 12k rows

    python eval/verify.py        # exits non-zero on any mismatch
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from eval.harness.runner import DEFAULT_LABELS, DEFAULT_WORKBOOKS, load_labels  # noqa
from eval.schemas import WorkbookLabel  # noqa: E402
from eval.util import range_box, safe_eval, values_match  # noqa: E402

_FOOT = re.compile(r"\s*\([a-z]\)\s*$")


def parse_raw(raw):
    """Recover the clean value from a cell's literal text."""
    if isinstance(raw, bool):
        return raw
    if isinstance(raw, (int, float)):
        return raw
    if isinstance(raw, str):
        t = _FOOT.sub("", raw.strip())
        neg = t.startswith("(") and t.endswith(")")
        if neg:
            t = t[1:-1]
        t = t.replace(",", "")
        try:
            v = float(t)
            v = int(v) if v.is_integer() else v
            return -v if neg else v
        except ValueError:
            return raw
    return raw


class Checker:
    def __init__(self):
        self.errors: list[str] = []
        self.checks = 0

    def expect(self, cond: bool, msg: str):
        self.checks += 1
        if not cond:
            self.errors.append(msg)


def verify_workbook(label: WorkbookLabel, wb_dir: Path, ck: Checker) -> None:
    wb = load_workbook(wb_dir / label.workbook, data_only=True)
    has_formula = any(c.is_formula for t in label.tables for c in t.cells)
    wbf = load_workbook(wb_dir / label.workbook, data_only=False) if has_formula else None
    cells_idx = {}

    for t in label.tables:
        ws = wb[t.sheet]
        r1, c1, r2, c2 = range_box(t.region)
        ck.expect(r2 <= ws.max_row + 0 and c2 <= ws.max_column,
                  f"{label.workbook}:{t.table_id} region {t.region} exceeds sheet "
                  f"({ws.max_row}x{ws.max_column})")

        for c in t.cells:
            cells_idx[(t.table_id, c.row_label, c.col_label)] = c.value
            if c.is_formula:
                # the cell really holds a formula...
                formula = wbf[t.sheet][c.cell_ref].value
                ck.expect(isinstance(formula, str) and formula.startswith("="),
                          f"{label.workbook}:{c.cell_ref} expected formula, got {formula!r}")
                # ...and its cached recalculated result equals the ground truth
                cached = ws[c.cell_ref].value
                ck.expect(cached is not None,
                          f"{label.workbook}:{c.cell_ref} formula has no cached value "
                          "(recalc step missing?)")
                ck.expect(values_match(c.value, cached, 1e-6, "number"),
                          f"{label.workbook}:{c.cell_ref} formula value {c.value!r} "
                          f"!= cached {cached!r}")
                continue
            actual = ws[c.cell_ref].value
            # literal text matches
            if isinstance(c.raw, (int, float)) and not isinstance(c.raw, bool):
                ck.expect(values_match(c.raw, actual, 1e-9, "number"),
                          f"{label.workbook}:{c.cell_ref} raw {c.raw!r} != {actual!r}")
            else:
                ck.expect(str(c.raw) == str(actual),
                          f"{label.workbook}:{c.cell_ref} raw {c.raw!r} != {actual!r}")
            # clean value recoverable from literal text
            recovered = parse_raw(actual)
            dtype = "number" if isinstance(c.value, (int, float)) else "string"
            ck.expect(values_match(c.value, recovered, 1e-9, dtype),
                      f"{label.workbook}:{c.cell_ref} value {c.value!r} not recoverable "
                      f"from {actual!r} (got {recovered!r})")

    for m in label.measures:
        key = (m.table_id, m.row_label, m.col_label)
        ck.expect(key in cells_idx,
                  f"{label.workbook}: measure {m.semantic_name} -> missing cell {key}")
        if key in cells_idx:
            dtype = "number" if isinstance(m.value, (int, float)) else "string"
            ck.expect(values_match(m.value, cells_idx[key], 1e-9, dtype),
                      f"{label.workbook}: measure {m.semantic_name} value {m.value} "
                      f"!= cell {cells_idx[key]}")

    region_by_id = {t.table_id: t.region for t in label.tables}
    measure_by_name = {m.semantic_name: m for m in label.measures}

    for s in label.samples:
        if s.type == "extraction":
            ws = wb[s.sheet]
            recovered = parse_raw(ws[s.expected_cell_ref].value)
            ck.expect(values_match(s.expected_value, recovered, s.tolerance, s.dtype),
                      f"{label.workbook}: ext {s.id} expected {s.expected_value} "
                      f"!= file {recovered}")
        elif s.type == "semantic":
            key = (s.expected_table_id, s.expected_row_label, s.expected_col_label)
            ck.expect(key in cells_idx and values_match(
                s.expected_value, cells_idx[key], s.tolerance, s.dtype),
                f"{label.workbook}: sem {s.id} expected {s.expected_value} != "
                f"cell {cells_idx.get(key)}")
        elif s.type == "boundary":
            ck.expect(region_by_id.get(s.table_id) == s.expected_region,
                      f"{label.workbook}: bnd {s.id} region mismatch")
        elif s.type == "formula":
            names = {}
            ok = True
            for sym, mname in s.operands.items():
                m = measure_by_name.get(mname)
                if not m:
                    ok = False
                    ck.expect(False, f"{label.workbook}: fml {s.id} operand {mname} "
                                     "not a known measure")
                    continue
                names[sym] = float(m.value)
                # operand input recorded must equal the measure value
                ck.expect(values_match(s.inputs[sym], m.value, 1e-9, "number"),
                          f"{label.workbook}: fml {s.id} input {sym}={s.inputs[sym]} "
                          f"!= measure {m.value}")
            if ok:
                recomputed = safe_eval(s.expression, names)
                ck.expect(values_match(s.expected_value, recomputed, s.tolerance,
                                       "number"),
                          f"{label.workbook}: fml {s.id} expected {s.expected_value} "
                          f"!= recomputed {recomputed}")

    if label.workbook == "large_ledger.xlsx":
        _verify_ledger_aggregates(label, wb, ck)


def _verify_ledger_aggregates(label: WorkbookLabel, wb, ck: Checker) -> None:
    """Recompute regional totals from all 12k ledger rows and check the summary."""
    ws = wb["Transactions"]
    totals_amt: dict[str, float] = {}
    totals_qty: dict[str, float] = {}
    # header in row 1: TxnID | Date | Region | Product | Amount | Qty
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        _, _, region, _, amount, qty = row
        totals_amt[region] = totals_amt.get(region, 0) + amount
        totals_qty[region] = totals_qty.get(region, 0) + qty
    summary = label.table("ledger_summary")
    by = {(c.row_label, c.col_label): c.value for c in summary.cells}
    for region, amt in totals_amt.items():
        ck.expect(values_match(by[(region, "TotalAmount")], amt, 1e-9, "number"),
                  f"large_ledger: {region} TotalAmount {by[(region,'TotalAmount')]} "
                  f"!= recomputed {amt}")
        ck.expect(values_match(by[(region, "TotalQty")], totals_qty[region], 1e-9,
                               "number"),
                  f"large_ledger: {region} TotalQty mismatch")


def verify_scale(label: WorkbookLabel, wb_dir: Path, ck: Checker) -> None:
    """Streaming verification for the extreme-scale workbook (read_only, one pass).

    Confirms: row count + width, deep spot-cells match labels, and the summary
    aggregates recompute exactly from the full ledger.
    """
    from eval.generator import scale_workbook as sw

    wb = load_workbook(wb_dir / label.workbook, data_only=True, read_only=True)
    cells_idx = {(t.table_id, c.row_label, c.col_label): c.value
                 for t in label.tables for c in t.cells}

    # cells we must spot-check, grouped by sheet -> {(row,col): value}
    needed: dict[str, dict[tuple[int, int], object]] = {}
    for t in label.tables:
        for c in t.cells:
            r, col = coordinate_to_tuple(c.cell_ref)
            needed.setdefault(t.sheet, {})[(r, col)] = c.value

    # ---- stream the giant ledger: aggregates + spot cells + dimensions ------
    ws = wb["Transactions"]
    header = None
    pos = {}
    agg = {"region": {}, "month": {}, "category": {}}
    n_data = 0
    width = 0
    need_tx = needed.get("Transactions", {})
    for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ridx == 1:
            header = list(row)
            width = len(header)
            pos = {name: i for i, name in enumerate(header)}
            continue
        if row[0] is None:
            continue
        n_data += 1
        region = row[pos["Region"]]
        month = str(row[pos["Date"]])[:7]
        cat = row[pos["ProductCategory"]]
        units = row[pos["UnitsSold"]]
        netrev = row[pos["NetRevenue"]]
        cogs = row[pos["COGS"]]
        gp = row[pos["GrossProfit"]]
        tax = row[pos["TaxAmount"]]
        for dim, key in (("region", region), ("month", month), ("category", cat)):
            d = agg[dim].setdefault(key, {"UnitsSold": 0, "NetRevenue": 0.0,
                                          "COGS": 0.0, "GrossProfit": 0.0,
                                          "TaxAmount": 0.0})
            d["UnitsSold"] += units
            d["NetRevenue"] += netrev
            d["COGS"] += cogs
            d["GrossProfit"] += gp
            d["TaxAmount"] += tax
        for (rr, cc), expected in need_tx.items():
            if rr == ridx:
                got = row[cc - 1]
                ck.expect(values_match(expected, got, 1e-9, "number"),
                          f"{label.workbook}: spot {get_col(cc)}{rr} {expected} != {got}")

    ck.expect(n_data == sw.N_ROWS,
              f"{label.workbook}: row count {n_data} != {sw.N_ROWS}")
    ck.expect(width == 1 + len(sw.COLUMNS),
              f"{label.workbook}: width {width} != {1 + len(sw.COLUMNS)}")

    # ---- spot-check the summary sheets --------------------------------------
    dim_of = {"region_summary": "region", "monthly_summary": "month",
              "category_summary": "category"}
    for t in label.tables:
        if t.sheet == "Transactions":
            continue
        wsx = wb[t.sheet]
        need = needed.get(t.sheet, {})
        for ridx, row in enumerate(wsx.iter_rows(values_only=True), start=1):
            for (rr, cc), expected in need.items():
                if rr == ridx:
                    got = row[cc - 1]
                    dtype = "number" if isinstance(expected, (int, float)) else "string"
                    ck.expect(values_match(expected, got, 1e-9, dtype),
                              f"{label.workbook}:{t.sheet} {get_col(cc)}{rr} "
                              f"{expected} != {got}")

    # ---- aggregates recomputed from ledger == summary cells ----------------
    for t in label.tables:
        dim = dim_of.get(t.table_id)
        if not dim:
            continue
        for c in t.cells:
            recomputed = agg[dim].get(c.row_label, {}).get(c.col_label)
            ck.expect(recomputed is not None and values_match(
                c.value, recomputed, 1e-6, "number"),
                f"{label.workbook}: {t.table_id} {c.row_label}/{c.col_label} "
                f"summary {c.value} != ledger-recomputed {recomputed}")

    # ---- samples ------------------------------------------------------------
    region_by_id = {t.table_id: t.region for t in label.tables}
    measure_by_name = {m.semantic_name: m for m in label.measures}
    for s in label.samples:
        if s.type == "extraction":
            ck.expect((s.table_id, s.row_label, s.col_label) in cells_idx and
                      values_match(s.expected_value,
                                   cells_idx[(s.table_id, s.row_label, s.col_label)],
                                   s.tolerance, s.dtype),
                      f"{label.workbook}: ext {s.id} mismatch")
        elif s.type == "semantic":
            key = (s.expected_table_id, s.expected_row_label, s.expected_col_label)
            ck.expect(key in cells_idx and values_match(
                s.expected_value, cells_idx[key], s.tolerance, s.dtype),
                f"{label.workbook}: sem {s.id} mismatch")
        elif s.type == "boundary":
            ck.expect(region_by_id.get(s.table_id) == s.expected_region,
                      f"{label.workbook}: bnd {s.id} region mismatch")
        elif s.type == "formula":
            names = {sym: float(measure_by_name[mn].value)
                     for sym, mn in s.operands.items()}
            ck.expect(values_match(s.expected_value, safe_eval(s.expression, names),
                                   s.tolerance, "number"),
                      f"{label.workbook}: fml {s.id} mismatch")


def get_col(idx: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def main() -> int:
    labels = load_labels(DEFAULT_LABELS)
    ck = Checker()
    for label in labels:
        is_scale = any("extreme_scale" in t.traps for t in label.tables)
        if is_scale:
            verify_scale(label, DEFAULT_WORKBOOKS, ck)
        else:
            verify_workbook(label, DEFAULT_WORKBOOKS, ck)
    print(f"Ran {ck.checks} ground-truth checks across {len(labels)} workbooks.")
    if ck.errors:
        print(f"\nFAILED: {len(ck.errors)} mismatches:")
        for e in ck.errors[:50]:
            print("  -", e)
        return 1
    print("ALL LABELS VERIFIED CORRECT ✓")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
