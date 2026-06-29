"""WorkbookSource port — abstracts where cells come from so the swarm depends on
an interface, not on openpyxl+path. Ships one impl (file); others (bytes, DataFrame,
streaming) implement the same Protocol later without touching extraction logic."""
from __future__ import annotations
from typing import Any, Optional, Protocol, runtime_checkable

import openpyxl


@runtime_checkable
class WorkbookSource(Protocol):
    def sheet_names(self) -> list[str]: ...
    def read_region(self, sheet: str, min_row: Optional[int] = None,
                    min_col: Optional[int] = None, max_row: Optional[int] = None,
                    max_col: Optional[int] = None) -> list[tuple]: ...
    def read_cell(self, sheet: str, row: int, col: int) -> Any: ...
    def read_formula_region(self, sheet: str, min_row: Optional[int] = None,
                            min_col: Optional[int] = None, max_row: Optional[int] = None,
                            max_col: Optional[int] = None) -> list[tuple]: ...


class OpenpyxlFileSource:
    """File-backed source. Opens read-only/data_only per call to preserve the existing
    live-read semantics (edits to a closed workbook are reflected on the next read)."""

    def __init__(self, path: str) -> None:
        self.path = path

    def sheet_names(self) -> list[str]:
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    def read_region(self, sheet, min_row=None, min_col=None, max_row=None, max_col=None):
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        try:
            ws = wb[sheet]
            return [r for r in ws.iter_rows(
                min_row=min_row, max_row=max_row,
                min_col=min_col, max_col=max_col, values_only=True)]
        finally:
            wb.close()

    def read_cell(self, sheet, row, col):
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        try:
            return wb[sheet].cell(row=row, column=col).value
        finally:
            wb.close()

    def read_formula_region(self, sheet, min_row=None, min_col=None, max_row=None, max_col=None):
        # data_only=False exposes formula strings; values_only=True avoids touching
        # cell.coordinate on EmptyCell objects (read-only mode crash, commit b77195b).
        wb = openpyxl.load_workbook(self.path, data_only=False, read_only=True)
        try:
            ws = wb[sheet]
            return [r for r in ws.iter_rows(
                min_row=min_row, max_row=max_row,
                min_col=min_col, max_col=max_col, values_only=True)]
        finally:
            wb.close()


class SnapshotSource:
    """Serves read_cell from an in-memory snapshot so callers that read cell-by-cell —
    e.g. ExtractionIndex.query() during the quality gate — avoid a workbook open per call.

    Backed by a {(row, col): value} dict for one sheet (typically the gate's already-read
    `live_cache`). read_region and any out-of-snapshot read_cell delegate to the wrapped
    source, so behaviour outside the snapshot window is unchanged. The snapshot is a
    point-in-time view: correct for the duration of one validation run."""

    def __init__(self, inner: WorkbookSource, sheet: str, cells: dict) -> None:
        self._inner = inner
        self._sheet = sheet
        self._cells = cells

    def sheet_names(self) -> list[str]:
        return self._inner.sheet_names()

    def read_region(self, sheet, min_row=None, min_col=None, max_row=None, max_col=None):
        return self._inner.read_region(sheet, min_row, min_col, max_row, max_col)

    def read_cell(self, sheet, row, col):
        if sheet == self._sheet and (row, col) in self._cells:
            return self._cells[(row, col)]
        return self._inner.read_cell(sheet, row, col)

    def read_formula_region(self, sheet, min_row=None, min_col=None, max_row=None, max_col=None):
        return self._inner.read_formula_region(sheet, min_row, min_col, max_row, max_col)


def as_source(x) -> WorkbookSource:
    """Normalize a path str, {'main': path} dict, or WorkbookSource into a WorkbookSource."""
    if isinstance(x, WorkbookSource) and not isinstance(x, (str, dict)):
        return x
    if isinstance(x, dict):
        return OpenpyxlFileSource(x["main"])
    if isinstance(x, str):
        return OpenpyxlFileSource(x)
    raise TypeError(f"cannot build WorkbookSource from {type(x).__name__}")
