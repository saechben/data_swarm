"""Deterministic coverage / residue detection.

Scans a sheet's full grid against the region(s) the splitter chose and flags
data or structure the static pass would otherwise silently corrupt or drop.
Model-free: this is the detection guarantee, independent of any agent.
"""
from __future__ import annotations

import re

from openpyxl.utils import get_column_letter, range_boundaries

from mcg_swarm.schemas import Finding
from mcg_swarm.splitter import _is_header_candidate

# accounting/currency/percent/number-ish text that signals a "header" row is really data
_VALUE_LIKE = re.compile(r"^\s*[\$€£]?\(?-?[\d.,]+\)?\s*%?\s*$")


def nonempty_cells(grid: list[tuple]) -> set[tuple[int, int]]:
    out: set[tuple[int, int]] = set()
    for i, row in enumerate(grid):
        for j, c in enumerate(row):
            if c not in (None, ""):
                out.add((i + 1, j + 1))
    return out


def region_cells(region: str) -> set[tuple[int, int]]:
    min_col, min_row, max_col, max_row = range_boundaries(region)
    return {(r, c) for r in range(min_row, max_row + 1)
            for c in range(min_col, max_col + 1)}


def coverage_score(grid: list[tuple], regions: list[str]) -> int:
    covered: set[tuple[int, int]] = set()
    for reg in regions:
        covered |= region_cells(reg)
    return len(nonempty_cells(grid) & covered)


def _components(cells: set[tuple[int, int]]) -> list[set[tuple[int, int]]]:
    """8-connected components over a set of (row, col) cells."""
    remaining = set(cells)
    comps: list[set[tuple[int, int]]] = []
    while remaining:
        seed = remaining.pop()
        stack = [seed]
        comp = {seed}
        while stack:
            r, c = stack.pop()
            for dr in (-1, 0, 1):
                for dc in (-1, 0, 1):
                    n = (r + dr, c + dc)
                    if n in remaining:
                        remaining.discard(n)
                        comp.add(n)
                        stack.append(n)
        comps.append(comp)
    return comps


def _subgrid(grid, minr, minc, maxr, maxc) -> list[tuple]:
    out = []
    for r in range(minr, maxr + 1):
        row = grid[r - 1] if r - 1 < len(grid) else ()
        out.append(tuple(row[c - 1] if c - 1 < len(row) else None
                         for c in range(minc, maxc + 1)))
    return out


def _a1(minr, minc, maxr, maxc) -> str:
    return f"{get_column_letter(minc)}{minr}:{get_column_letter(maxc)}{maxr}"


def scan_handle(grid: list[tuple], handle, sheet: str) -> list[Finding]:
    """Deterministic detection over one sheet's grid vs its chosen handle region."""
    findings: list[Finding] = []
    try:
        # ---- uncovered-data: nonempty blocks outside the region that look tabular ----
        nonempty = nonempty_cells(grid)
        covered = region_cells(handle.region)
        uncovered = nonempty - covered
        for comp in _components(uncovered):
            minr = min(r for r, _ in comp)
            maxr = max(r for r, _ in comp)
            minc = min(c for _, c in comp)
            maxc = max(c for _, c in comp)
            sub = _subgrid(grid, minr, minc, maxr, maxc)
            if sub and _is_header_candidate(sub[0], sub[1:]):
                findings.append(Finding(
                    category="uncovered-data", severity="error", scope="sheet",
                    source="static", ref=f"{sheet}!{_a1(minr, minc, maxr, maxc)}",
                    message=(f"uncovered tabular block at {sheet}!"
                             f"{_a1(minr, minc, maxr, maxc)} outside detected region "
                             f"{handle.region} — a second table was likely dropped")))

        # ---- header-row inspection (empty corner / false span / transpose) ----
        min_col, min_row, max_col, max_row = range_boundaries(handle.region)
        hr = handle.header_row
        header = _subgrid(grid, hr, min_col, hr, max_col)[0] if hr - 1 < len(grid) else ()
        if header and header[0] in (None, ""):
            findings.append(Finding(
                category="empty-header-corner", severity="error", scope="table",
                source="static", ref=f"{sheet}!{get_column_letter(min_col)}{hr}",
                message=(f"empty top-left header cell at {sheet}!"
                         f"{get_column_letter(min_col)}{hr} — header/orientation is ambiguous "
                         f"(transposed or corner-labelled table)")))

        span = getattr(handle, "header_span", 1)
        if span >= 2 and hr < len(grid):
            leaf = _subgrid(grid, hr + 1, min_col, hr + 1, max_col)[0]
            leaf_vals = [c for c in leaf if c not in (None, "")]
            value_like = [c for c in leaf_vals
                          if isinstance(c, (int, float))
                          or (isinstance(c, str) and _VALUE_LIKE.match(c))]
            if leaf_vals and len(value_like) >= max(1, len(leaf_vals) // 2):
                findings.append(Finding(
                    category="false-header-span", severity="error", scope="table",
                    source="static", ref=f"{sheet}!{handle.region}",
                    message=(f"header_span=2 but row {hr + 1} looks like data "
                             f"(value-like cells) — first data row likely consumed as a header")))

        # ---- transpose-suspected: empty corner + left col labels + top row labels ----
        if header and header[0] in (None, ""):
            below_first = [grid[r - 1][min_col - 1]
                           for r in range(hr + 1, min(max_row, len(grid)) + 1)
                           if min_col - 1 < len(grid[r - 1])]
            top_after = [c for c in header[1:] if c not in (None, "")]
            below_str = [c for c in below_first if c not in (None, "")]
            if (below_str and all(isinstance(c, str) for c in below_str)
                    and top_after and all(isinstance(c, str) for c in top_after)):
                findings.append(Finding(
                    category="transpose-suspected", severity="warning", scope="table",
                    source="static", ref=f"{sheet}!{handle.region}",
                    message=(f"sheet {sheet} may be transposed (labels down column "
                             f"{get_column_letter(min_col)}, periods across row {hr})")))
    except Exception:
        return findings  # never raise — detection is best-effort-safe, deterministic
    return findings
