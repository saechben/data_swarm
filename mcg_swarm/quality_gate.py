# mcg_swarm/quality_gate.py
"""In-loop quality gate: coverage (resolution-only) + index integrity + round-trip + computed-column checks."""
from __future__ import annotations

import datetime as _dt
from dataclasses import dataclass, field

from openpyxl.utils import coordinate_to_tuple, get_column_letter

from eval.util import range_box, values_match
from mcg_swarm.formulas import build_env, evaluate
from mcg_swarm.source import as_source, SnapshotSource

DTYPE_MISMATCH_TOL = 0.2  # > this fraction of non-null sampled cells off-type => fail


def _conforms(value, dtype: str) -> bool:
    """Return True if *value* is consistent with the declared *dtype*."""
    if value in (None, ""):
        return True
    if dtype == "string":
        return True
    if dtype == "boolean":
        return isinstance(value, bool)
    if dtype == "date":
        return isinstance(value, _dt.datetime)
    if dtype == "number":
        if isinstance(value, bool):
            return False
        if isinstance(value, (int, float)):
            return True
        try:
            float(str(value).replace(",", "").strip())
            return True
        except (TypeError, ValueError):
            return False
    return True


@dataclass
class TableTestReport:
    passed: bool
    failures: list = field(default_factory=list)




def run_table_tests(source, table, index, sample_size: int = 25) -> TableTestReport:
    """
    Run deterministic in-loop quality checks on an extracted table.

    Five phases (plus a reverse-coverage check folded into phase 2a):
    1. Coverage (resolution-only, O(keys+cols), zero file I/O):
       Every column in table.columns must be resolvable via the index
       (present in _col_to_phys) — a schema-vs-index coverage check, not a
       tautological comparison of the index against itself.  Also flags
       row-key collisions (index.duplicate_row_keys: an earlier row shadowed
       by a later one with the same key), blank-key rows (index.blank_key_rows:
       a data row with an empty key cell, unreachable by any meaningful key),
       and an empty index (zero row keys resolved — the table cannot be
       queried at all).  No per-cell reads.

    2. Index integrity (one workbook open total, O(cols) + O(sample) reads):
       (a) Column-name gate: reads the live header row(s) and asserts every
           declared column name is present, flags duplicate declared names,
           and — the reverse direction — flags any live header in the region
           that is NOT declared in table.columns (a silently dropped column).
       (b) Column-integrity: asserts name->physical_col matches _col_to_phys
           for each column.  Catches numeric->numeric column remaps.
       (c) Row-integrity: for each sampled key, reads the key-column cell at the
           index-resolved physical row and asserts it equals the key.

    3. Round-trip (bounded sample, ≤ sample_size × cols file reads):
       For a deterministic subset of row keys, compare index.query() value against
       the live cell value read directly from the workbook.

    4. Computed columns (same sample):
       For columns with role="computed" and a matching TableFormula, re-evaluate
       the formula and compare against the live cell.

    5. Dtype conformance (same sample, via the real index.query() path):
       For non-key, non-string columns, re-resolve each sampled cell through
       index.query() (backed by a snapshot of the already-read cells, so no
       extra workbook opens) and assert the value conforms to the column's
       declared dtype, failing only when the non-conforming fraction exceeds
       DTYPE_MISMATCH_TOL on a large-enough sample.
    """
    failures: list[str] = []

    keys = index.row_keys()
    cols = index.column_names()

    # ------------------------------------------------------------------
    # Phase 1: REAL coverage — schema vs index, and row-resolution integrity.
    # (The previous version compared index.column_names()/row_keys() against
    # the very dicts they are read from — a tautology that could never fail.)
    # ------------------------------------------------------------------
    for col in table.columns:
        if col.name not in index._col_to_phys:
            failures.append(
                f"coverage gap: column {col.name!r} in table.columns but not "
                "resolvable via the index")
    for key, shadowed, winner in getattr(index, "duplicate_row_keys", []):
        failures.append(
            f"row-key collision: key {key!r} at row {shadowed} is shadowed by "
            f"row {winner} — the earlier row is unreachable via query()")
    for r in getattr(index, "blank_key_rows", []):
        failures.append(
            f"blank row key: data row {r} has an empty key cell — the row is "
            "not reachable by a meaningful key")
    if not keys:
        failures.append(
            "empty index: zero row keys resolved — the table cannot be queried")

    # ------------------------------------------------------------------
    # Build deterministic sample (used by phases 2b, 3, 4).
    # Spread sample (head/middle/tail) so anomalies anywhere are caught, not just the
    # first rows. An explicit non-default sample_size still caps it for small-sample callers.
    # ------------------------------------------------------------------
    from mcg_swarm.sampling import select_sample
    sample_keys = select_sample(keys, sample_size=(sample_size if sample_size != 25 else None))

    # ------------------------------------------------------------------
    # Phases 2 + 3 combined: ONE workbook open for all live-file checks.
    # Replaces N*cols separate _live_value() opens (each ~2-3 s on large files).
    # ------------------------------------------------------------------
    min_row, min_col, max_row, max_col = range_box(table.region)
    header_row = table.header_row
    header_span = getattr(table, "header_span", 1)

    # Build all (phys_row, phys_col) cells we need to read:
    # Phase 2a (col-name gate): ALL header rows in span, all cols in region
    # Phase 2b (col-integrity): same header rows
    # Phase 2c (row-integrity): key-column cells for each sample key
    # Phase 3  (round-trip):    every (sample_key, col) cell

    row_key_names = table.extraction.row_key
    key_col_name = row_key_names[0] if row_key_names else None
    key_phys_col = index._col_to_phys.get(key_col_name) if key_col_name else None

    # collect all (r, c) pairs we need; read in one pass
    needed: set[tuple[int, int]] = set()
    # all header rows in the span
    for hr in range(header_row, header_row + header_span):
        for c in range(min_col, max_col + 1):
            needed.add((hr, c))
    # row-integrity: key-column cells
    if key_phys_col is not None:
        for k in sample_keys:
            if k in index._key_to_phys:
                needed.add((index._key_to_phys[k], key_phys_col))
    # round-trip: derive cell locations directly from index maps (no query() opens).
    # index.query() opens the workbook per call — O(sample*cols) opens on large files.
    # Instead build (phys_row, phys_col) from _key_to_phys/_col_to_phys and read once.
    sample_cells: dict[tuple, tuple[int, int, str]] = {}  # (k,col) -> (pr, pc, cell_ref)
    for k in sample_keys:
        if k not in index._key_to_phys:
            continue
        pr = index._key_to_phys[k]
        for col in cols:
            if col not in index._col_to_phys:
                continue
            pc = index._col_to_phys[col]
            cell_ref = f"{get_column_letter(pc)}{pr}"
            sample_cells[(k, col)] = (pr, pc, cell_ref)
            needed.add((pr, pc))

    # ONE workbook open — scan only the bounding box of sample rows.
    # sample_keys come from select_sample (head/middle/tail spread), so on large tables
    # the scan bounding box can span most of the sheet height — see OPTIMIZATIONS.md #1.
    # (rows header_row .. max(sample_phys_rows), all cols in region).
    live_cache: dict[tuple[int, int], object] = {}
    if needed:
        need_rows = [r for (r, _) in needed]
        need_cols_list = [c for (_, c) in needed]
        scan_min_row = min(need_rows)
        scan_max_row = max(need_rows)
        scan_min_col = min(need_cols_list)
        scan_max_col = max(need_cols_list)
        src = as_source(source)
        rows = src.read_region(table.sheet, scan_min_row, scan_min_col,
                               scan_max_row, scan_max_col)
        for r_off, row_vals in enumerate(rows):
            actual_row = scan_min_row + r_off
            for c_off, val in enumerate(row_vals):
                pos = (actual_row, scan_min_col + c_off)
                if pos in needed:
                    live_cache[pos] = val

    # ── Phase 5: dtype conformance (validated through the REAL query() path) ─
    # The structural checks above never validate that a column's VALUES match its
    # declared dtype. Crucially, we validate the value the agent will actually receive:
    # each sampled cell is read via index.query() — the same function downstream consumers
    # call — not the raw batch read. query() opens the file per cell (O(rows) in read-only
    # mode), so for the bounded sample we point it at a SnapshotSource backed by live_cache:
    # one read, query()'s real resolution + dtype tagging exercised, and no per-cell
    # workbook opens (the OPT-2 large-file hang). The independent fresh-read oracle that
    # proves snapshot == live cell stays in Phase 3 round-trip.
    key_names = set(table.extraction.row_key or [])
    _orig_src = index.source
    index.source = SnapshotSource(_orig_src, index.sheet, live_cache)
    try:
        for col in table.columns:
            if col.name in key_names or col.dtype == "string":
                continue
            if col.name not in index._col_to_phys:
                continue
            total = bad = 0
            for k in sample_keys:
                if k not in index._key_to_phys:
                    continue
                try:
                    val = index.query(k, col.name).value
                except Exception:
                    continue
                if val in (None, ""):
                    continue
                total += 1
                if not _conforms(val, col.dtype):
                    bad += 1
            if total >= 5 and (bad / total) > DTYPE_MISMATCH_TOL:
                failures.append(
                    f"dtype-mismatch: column {col.name!r} declared {col.dtype} but "
                    f"{bad}/{total} sampled non-null cells are not {col.dtype}"
                )
    finally:
        index.source = _orig_src

    # ── Phase 2a: column-name gate ──────────────────────────────────────
    # Build composite live_col_map from all header_span rows using bottom-first rule.
    live_col_map: dict[str, int] = {}
    for c in range(min_col, max_col + 1):
        name = None
        # scan bottom header row → top
        for hr in range(header_row + header_span - 1, header_row - 1, -1):
            val = live_cache.get((hr, c))
            if val not in (None, ""):
                name = str(val)
                break
        if name is not None:
            live_col_map[name] = c
    live_names: set[str] = set(live_col_map.keys())

    col_names_list = [c.name for c in table.columns]
    seen_names: set[str] = set()
    for name in col_names_list:
        if name in seen_names:
            failures.append(f"column-name: duplicate column name {name!r} in table.columns")
        seen_names.add(name)
    for name in col_names_list:
        if name not in live_names:
            failures.append(
                f"column-name: {name!r} in table.columns not found in live header rows "
                f"{header_row}–{header_row + header_span - 1}"
            )

    # Reverse coverage: every live header in the region must be declared —
    # a dropped column is silent data loss for every downstream consumer.
    declared = {c.name for c in table.columns}
    for name in live_names:
        if name not in declared:
            failures.append(
                f"column-coverage: live header {name!r} in region not declared "
                "in table.columns — column silently dropped")

    # ── Phase 2b: column-integrity ──────────────────────────────────────
    for col_name, idx_phys in index._col_to_phys.items():
        live_phys = live_col_map.get(col_name)
        if live_phys is None:
            failures.append(
                f"column-integrity: {col_name!r} in index but not found in live header rows "
                f"{header_row}–{header_row + header_span - 1}"
            )
        elif idx_phys != live_phys:
            failures.append(
                f"column-integrity: {col_name!r} index col={get_column_letter(idx_phys)} "
                f"but live header says col={get_column_letter(live_phys)}"
            )

    # ── Phase 2c: row-integrity ─────────────────────────────────────────
    if key_phys_col is not None:
        for k in sample_keys:
            if k not in index._key_to_phys:
                continue
            phys_row = index._key_to_phys[k]
            live_key = live_cache.get((phys_row, key_phys_col))
            if live_key != k:
                failures.append(
                    f"row-integrity: key {k!r} -> row {phys_row} "
                    f"but live cell {get_column_letter(key_phys_col)}{phys_row}={live_key!r}"
                )

    # ── Phase 3: round-trip (REAL) ──────────────────────────────────────
    # For a small subsample (≤ ROUND_TRIP_SUBSAMPLE keys), call index.query(k, col)
    # which reopens the live file at the index-resolved address, then compare the
    # returned value against the independently-read live_cache cell.
    # This catches divergence between what query() returns and the live batch read.
    # Bounded to ~5×cols file opens so it stays cheap even on large workbooks.
    ROUND_TRIP_SUBSAMPLE = 5
    rt_keys = sample_keys[:ROUND_TRIP_SUBSAMPLE]
    for k in rt_keys:
        for col in cols:
            info = sample_cells.get((k, col))
            if info is None:
                continue
            pr, pc, cell_ref = info
            live = live_cache.get((pr, pc))
            spec = index.columns.get(col)
            dtype = "number" if (spec and spec.dtype == "number") else "string"
            try:
                queried = index.query(k, col).value
            except Exception as e:
                failures.append(f"round-trip: query({k!r}, {col!r}) raised {e}")
                continue
            # Both None → cell is genuinely empty and index agrees; not a failure.
            if live is None and queried is None:
                continue
            # Round-trip is a CONSISTENCY check between two reads of the same cell, so
            # identical values always match — the dtype/tolerance compare only matters
            # when query() normalizes representation (e.g. "1,200" -> 1200.0). Without
            # this, a number column holding a non-numeric missing-value sentinel ('n/a')
            # would float-coerce-fail against its own identical live read (a false
            # positive that punishes correctly tightening a dtype to number).
            if live == queried or str(live).strip() == str(queried).strip():
                continue
            if not values_match(live, queried, 1e-9, dtype):
                failures.append(
                    f"round-trip: {col!r}@{k!r} live={live!r} but query()={queried!r}"
                )

    # ------------------------------------------------------------------
    # Phase 4: Computed columns — same sample (no index.query() opens needed
    # because computed columns are rare; skip on tables with no formulas).
    # ------------------------------------------------------------------
    formulas_by_target = {f.target: f for f in table.formulas}
    computed_cols = [c for c in table.columns
                     if c.role == "computed" and c.name in formulas_by_target]
    if computed_cols:
        for c in computed_cols:
            f = formulas_by_target[c.name]
            for k in sample_keys:
                try:
                    got = evaluate(
                        f,
                        build_env(
                            f,
                            k,
                            index.query,
                            query_cell=index.query_cell,
                            query_range=index.query_range,
                        ),
                    )
                    info = sample_cells.get((k, c.name))
                    live = live_cache.get((info[0], info[1])) if info else None
                    if not values_match(live, got, 1e-6, "number"):
                        failures.append(
                            f"computed mismatch {c.name}@{k}: live={live} calc={got}"
                        )
                except Exception as e:
                    failures.append(f"computed eval failed {c.name}@{k}: {e}")

    return TableTestReport(passed=not failures, failures=failures)
