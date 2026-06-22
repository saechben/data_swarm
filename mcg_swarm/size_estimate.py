from __future__ import annotations
from dataclasses import dataclass
from math import ceil
from openpyxl.utils import get_column_letter
from eval.util import range_box

ROWS_PER_AGENT = 5_000
COLS_PER_AGENT = 40
K_MAX = 4  # cap bands to avoid O(K) workbook opens on huge files (each open ~2-3 s)


@dataclass
class SizeEstimate:
    rows: int
    cols: int
    cell_count: int


@dataclass
class Band:
    sheet: str
    header_row: int
    region: str
    col_start: int
    col_end: int
    row_start: int
    row_end: int


def size_estimate(handle) -> SizeEstimate:
    min_row, min_col, max_row, max_col = range_box(handle.region)
    rows = max_row - handle.header_row          # data rows below header
    cols = max_col - min_col + 1
    return SizeEstimate(rows, cols, rows * cols)


def plan_bands(handle) -> tuple[str, int, list[Band]]:
    se = size_estimate(handle)
    min_row, min_col, max_row, max_col = range_box(handle.region)
    data_start = handle.header_row + 1
    row_pressure = se.rows / ROWS_PER_AGENT
    col_pressure = se.cols / COLS_PER_AGENT
    if max(row_pressure, col_pressure) <= 1:
        axis, k = "row", 1
    elif row_pressure >= col_pressure:
        axis, k = "row", min(ceil(row_pressure), K_MAX)
    else:
        axis, k = "col", min(ceil(col_pressure), K_MAX)

    bands: list[Band] = []
    if axis == "row":
        step = ceil(se.rows / k)
        for b in range(k):
            rs = data_start + b * step
            re = min(max_row, rs + step - 1)
            if rs > max_row:
                break
            region = (
                f"{get_column_letter(min_col)}{handle.header_row}"
                f":{get_column_letter(max_col)}{re}"
            )
            bands.append(Band(handle.sheet, handle.header_row, region,
                              min_col, max_col, rs, re))
    else:
        step = ceil(se.cols / k)
        for b in range(k):
            cs = min_col + b * step
            ce = min(max_col, cs + step - 1)
            if cs > max_col:
                break
            region = (
                f"{get_column_letter(cs)}{handle.header_row}"
                f":{get_column_letter(ce)}{max_row}"
            )
            bands.append(Band(handle.sheet, handle.header_row, region,
                              cs, ce, data_start, max_row))

    return axis, len(bands), bands
