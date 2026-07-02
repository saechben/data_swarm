from __future__ import annotations
import datetime as _dt
from dataclasses import dataclass, field
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from mcg_swarm.schemas import ColumnSpec


@dataclass
class TableHandle:
    sheet: str
    region: str
    header_row: int
    columns: list[ColumnSpec] = field(default_factory=list)
    ambiguous: bool = False
    reason: str = ""
    header_span: int = 1


def _infer_dtype(samples: list) -> str:
    vals = [v for v in samples if v not in (None, "")]
    if not vals:
        return "string"
    if all(isinstance(v, bool) for v in vals):
        return "boolean"
    if all(isinstance(v, _dt.datetime) for v in vals):
        return "date"

    def isnum(v):
        try:
            float(v)
            return not isinstance(v, bool)
        except (TypeError, ValueError):
            return False

    if all(isnum(v) for v in vals):
        return "number"
    return "string"


def _is_header_candidate(row: tuple, rows_after: list[tuple]) -> bool:
    """Return True if this row looks like a real table header (mostly strings, followed immediately by a data row)."""
    nonempty = [c for c in row if c not in (None, "")]
    if not nonempty:
        return False
    strings = [c for c in nonempty if isinstance(c, str)]
    if len(strings) < max(1, len(nonempty) // 2):
        return False
    # Require that the very next non-empty row exists (data must follow directly or after a gap,
    # but a single-cell row with a blank separator below is likely a title banner, not a header)
    has_data_after = any(any(c not in (None, "") for c in r) for r in rows_after)
    if not has_data_after:
        return False
    # If this candidate has only 1 non-empty cell, it is a title banner if:
    #   (a) the very next row is blank (existing case), OR
    #   (b) the very next row has strictly MORE non-empty cells (merged-title banner directly
    #       above the real header — openpyxl gives None for the non-anchor merged cells).
    if len(nonempty) == 1 and rows_after:
        next_row_nonempty = [c for c in rows_after[0] if c not in (None, "")]
        if all(c in (None, "") for c in rows_after[0]):
            # case (a): blank separator row below
            return False
        if len(next_row_nonempty) > 1:
            # case (b): wider row immediately below — this row is a title banner
            return False
    return True


def _is_title_banner(row: tuple, rows_after: list[tuple]) -> bool:
    """Return True if this row is a title banner that should be skipped (not ambiguous content)."""
    nonempty = [c for c in row if c not in (None, "")]
    if len(nonempty) != 1:
        return False
    if not rows_after:
        return False
    # Case (a): single cell followed by a blank row
    if all(c in (None, "") for c in rows_after[0]):
        return True
    # Case (b): single cell immediately above a wider row (merged title banner)
    next_nonempty = [c for c in rows_after[0] if c not in (None, "")]
    if len(next_nonempty) > 1:
        return True
    return False


def _is_secondary_header(row: tuple, rows_after: list[tuple]) -> bool:
    """Return True if *row* looks like a second (leaf) header row immediately below a group header.

    The distinction from a data row:
    - A secondary header is ALL-STRINGS: every non-empty cell is a string (no numerics).
      Data rows have at least one numeric value column.
    - There must be at least 2 non-empty string cells (a single-string row could be a key cell).
    - The row immediately after must be data (has at least one numeric value).
    """
    nonempty = [c for c in row if c not in (None, "")]
    if not nonempty:
        return False

    def _isnum(v):
        try:
            float(v)
            return not isinstance(v, bool)
        except (TypeError, ValueError):
            return False

    # A secondary header row must be PURE LABELS — every non-empty cell is a string
    # (no numerics, no booleans, no dates). Data rows always carry a non-string value
    # column; this rejects e.g. ("Status", True, "Active") which has 2 strings + a bool.
    if not all(isinstance(c, str) for c in nonempty):
        return False
    strings = nonempty
    # Must have at least 2 string cells (distinguishes data rows that have only 1 string key)
    if len(strings) < 2:
        return False
    # The very next row must be data (has at least one numeric value)
    if not rows_after:
        return False
    next_row = rows_after[0]
    next_nonempty = [c for c in next_row if c not in (None, "")]
    if not next_nonempty:
        return False
    nums_in_next = [v for v in next_nonempty if _isnum(v)]
    return len(nums_in_next) >= 1


def _composite_col_names(header_rows: list[tuple], first_col: int, last_col: int) -> list[str]:
    """Build composite column names using the bottom-row-first, scan-upward rule.

    For each physical column in [first_col, last_col] (0-based indices into the row tuples):
      name = bottom header row value if non-empty, else scan upward to first non-empty.
    """
    names = []
    n_hdr = len(header_rows)
    for j in range(first_col, last_col + 1):
        name = None
        # scan from bottom header row upward
        for row_idx in range(n_hdr - 1, -1, -1):
            row = header_rows[row_idx]
            val = row[j] if j < len(row) else None
            if val not in (None, ""):
                name = str(val)
                break
        if name is None:
            name = get_column_letter(j + 1)
        names.append(name)
    return names


def detect_table(rows: list[tuple], sheet_name: str) -> TableHandle:
    # find first non-empty row that is mostly strings and has a data row after it
    header_idx = None
    banner_rows: set[int] = set()
    for i, row in enumerate(rows):
        # Track title banner rows so they don't trigger "ambiguous" later
        if _is_title_banner(row, rows[i + 1:]):
            banner_rows.add(i)
            continue
        if _is_header_candidate(row, rows[i + 1:]):
            header_idx = i
            break
    if header_idx is None:
        return TableHandle(sheet_name, "A1:A1", 1, [], ambiguous=True,
                           reason="no header row with data below")
    header = rows[header_idx]

    # PATTERN C: detect 2-row header (secondary/leaf row immediately below top header).
    # Cap at span=2; no deeper headers in scope.
    rows_after_header = rows[header_idx + 1:]
    header_span = 1
    if rows_after_header and _is_secondary_header(rows_after_header[0], rows_after_header[1:]):
        header_span = 2

    # All header rows as a list (1 or 2 rows)
    header_rows = rows[header_idx: header_idx + header_span]

    # PATTERN A: walk upward to include contiguous banner rows above the header
    # so the region top captures title banners (e.g. capex row2 title + row3 header).
    top_idx = header_idx
    while top_idx - 1 >= 0 and (top_idx - 1) in banner_rows:
        top_idx -= 1

    # PATTERN B: right-edge stops at first fully-empty gap column (contiguous run).
    # For multi-row headers, use the UNION of non-empty cells across ALL header rows
    # (so sparse group-header rows extend the right edge to match their leaf rows).
    # Provisional last_col = rightmost non-empty cell across all header rows.
    provisional_last = max(
        (j for hrow in header_rows for j, c in enumerate(hrow) if c not in (None, "")),
        default=0
    )
    # Data rows start AFTER the last header row.
    data_start_idx = header_idx + header_span
    provisional_end = data_start_idx - 1
    for r in range(data_start_idx, len(rows)):
        if all(c in (None, "") for c in rows[r][: provisional_last + 1]):
            break
        provisional_end = r
    data_provisional = rows[data_start_idx: provisional_end + 1]

    # find first_col (leftmost non-empty in any header row or data)
    all_rows_provisional = list(header_rows) + list(data_provisional)
    first_col = 0
    for j in range(provisional_last + 1):
        if any(j < len(r) and r[j] not in (None, "") for r in all_rows_provisional):
            first_col = j
            break

    # PATTERN B: contiguous right-edge run starting from first_col.
    # Stop at the first column that is empty in BOTH all header rows AND all data rows.
    last_col = first_col
    max_width = max(len(r) for r in all_rows_provisional) if all_rows_provisional else first_col + 1
    for j in range(first_col, max_width):
        col_is_empty = (
            all(j >= len(hrow) or hrow[j] in (None, "") for hrow in header_rows)
            and all(j >= len(r) or r[j] in (None, "") for r in data_provisional)
        )
        if col_is_empty:
            break
        last_col = j

    # Re-derive end_idx now that last_col is correct (gap-aware).
    end_idx = data_start_idx - 1
    for r in range(data_start_idx, len(rows)):
        if all(c in (None, "") for c in rows[r][: last_col + 1]):
            break
        end_idx = r

    data = rows[data_start_idx: end_idx + 1]
    all_table_rows = list(header_rows) + list(data)

    # last_col_trimmed: trim trailing empty cols within [first_col, last_col]
    last_col_trimmed = last_col
    for j in range(last_col, first_col - 1, -1):
        if any(j < len(r) and r[j] not in (None, "") for r in all_table_rows):
            last_col_trimmed = j
            break

    start_col_letter = get_column_letter(first_col + 1)
    end_col_letter = get_column_letter(last_col_trimmed + 1)
    # PATTERN A: region starts at top_idx (includes banner rows), header stays at header_idx+1
    region = f"{start_col_letter}{top_idx + 1}:{end_col_letter}{end_idx + 1}"

    # PATTERN C: composite column names across header span (bottom-first, scan upward)
    col_names = _composite_col_names(header_rows, first_col, last_col_trimmed)
    cols = []
    for idx_in_names, j in enumerate(range(first_col, last_col_trimmed + 1)):
        name = col_names[idx_in_names]
        samples = [r[j] if j < len(r) else None for r in data[:20]]
        cols.append(ColumnSpec(name=name, dtype=_infer_dtype(samples),
                               role="key" if j == first_col else "value"))

    # Non-empty rows above the header that are NOT recognised title banners make it ambiguous.
    ambiguous = header_idx > 0 and any(
        k not in banner_rows and any(c not in (None, "") for c in rows[k])
        for k in range(header_idx))
    return TableHandle(sheet_name, region, header_idx + 1, cols,
                       ambiguous=ambiguous,
                       reason="content above header row" if ambiguous else "",
                       header_span=header_span)


def split_workbook(source, config=None) -> list[TableHandle]:
    """Split a workbook into TableHandles via the active analyzer lenses.

    Back-compat shim over ``analyze_workbook()``: flattens the per-sheet winning
    handles and drops per-sheet view/method/findings. Rich callers (run_swarm)
    use ``mcg_swarm.analyzers.pipeline.analyze_workbook`` directly.
    """
    # Lazy import: analyzers import TableHandle/detect_table from this module.
    from mcg_swarm.analyzers.pipeline import analyze_workbook
    return [h for sa in analyze_workbook(source, config) for h in sa.handles]


def handle_from_region(grid: list[tuple], sheet: str, region: str,
                       header_row: int, header_span: int = 1) -> TableHandle:
    """Build a TableHandle honouring an explicit absolute region + header row/span.

    Used to materialise an agent's re-cut proposal into a real handle. Column names come
    from the header span (bottom-row-first composite rule); dtypes are inferred from the
    data rows below the header. `header_row` is a 1-based absolute sheet row.
    """
    min_col, min_row, max_col, max_row = range_boundaries(region)

    def cell(r: int, c: int):
        row = grid[r - 1] if 0 <= r - 1 < len(grid) else ()
        return row[c - 1] if 0 <= c - 1 < len(row) else None

    header_rows = [
        tuple(cell(header_row + k, c) for c in range(min_col, max_col + 1))
        for k in range(header_span)
    ]
    data_rows = [
        tuple(cell(r, c) for c in range(min_col, max_col + 1))
        for r in range(header_row + header_span, max_row + 1)
    ]
    names = _composite_col_names(header_rows, 0, max_col - min_col)
    cols = []
    for j in range(max_col - min_col + 1):
        samples = [dr[j] if j < len(dr) else None for dr in data_rows[:20]]
        cols.append(ColumnSpec(name=names[j], dtype=_infer_dtype(samples),
                               role="key" if j == 0 else "value"))
    return TableHandle(sheet, region, header_row, cols, header_span=header_span)
