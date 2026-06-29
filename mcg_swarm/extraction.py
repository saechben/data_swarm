# mcg_swarm/extraction.py
from __future__ import annotations
import logging
from openpyxl.utils import get_column_letter
from eval.util import range_box
from mcg_swarm.schemas import ExtractedValue

_log = logging.getLogger(__name__)


def _composite_col_map(grid: list, hdr_off: int, header_span: int,
                       min_col: int) -> dict:
    """Build column-name -> absolute-physical-col map using composite (bottom-first) rule.

    For each column position, the name is the bottom header row value if non-empty,
    else scan upward to the first non-empty cell across the header span.
    Returns {col_name: abs_phys_col}.
    """
    if header_span < 1:
        header_span = 1
    header_rows = grid[hdr_off: hdr_off + header_span]
    if not header_rows:
        return {}
    n_cols = max(len(r) for r in header_rows) if header_rows else 0
    col_map: dict[str, int] = {}
    for j in range(n_cols):
        name = None
        # scan bottom → top
        for row_idx in range(len(header_rows) - 1, -1, -1):
            row = header_rows[row_idx]
            val = row[j] if j < len(row) else None
            if val not in (None, ""):
                name = str(val)
                break
        if name is not None:
            abs_col = min_col + j
            if name in col_map:
                # Duplicate header name: the later column overwrites the earlier,
                # making the earlier column unreachable via query() and silently
                # shifting which physical cell `name` resolves to. Surface it.
                _log.warning(
                    "duplicate header name %r: column %s overwrites column %s "
                    "(earlier column becomes unreachable via query())",
                    name, get_column_letter(abs_col),
                    get_column_letter(col_map[name]),
                )
            col_map[name] = abs_col
    return col_map


class ExtractionIndex:
    def __init__(self, source, sheet, region, header_row, columns, row_key,
                 header_span: int = 1):
        self.source, self.sheet = source, sheet
        self.columns = {c.name: c for c in columns}
        self.row_key = row_key
        min_row, min_col, max_row, max_col = range_box(region)
        self._min_col = min_col

        # Build-time scan: read region once to precompute dicts (O(1) resolution).
        # Live reads use source.read_cell per call (correctness over speed).
        grid = self.source.read_region(sheet, min_row, min_col, max_row, max_col)

        # PATTERN A: header may not be grid[0] when region top is above the header
        # (i.e. region includes title-banner rows). Use header_row to locate the
        # correct grid offset so hdr_off==0 in the normal case (no change in behaviour).
        hdr_off = header_row - min_row          # 0-based offset of header within grid

        # PATTERN C: composite column map across header_span rows (bottom-first).
        # For header_span==1 this is identical to reading grid[hdr_off] directly.
        self._col_to_phys: dict[str, int] = _composite_col_map(
            grid, hdr_off, header_span, min_col
        )

        # row key value -> absolute (1-based) row index
        # Data rows start AFTER all header rows (hdr_off + header_span).
        data_start_off = hdr_off + header_span
        data_start_row = header_row + header_span  # absolute 1-based

        self._key_to_phys: dict = {}
        key_cols = [self._col_to_phys[k] for k in row_key] if row_key else []
        for i, row in enumerate(grid[data_start_off:], start=data_start_row):
            if row_key:
                vals = tuple(row[kc - min_col] for kc in key_cols)
                key = vals[0] if len(vals) == 1 else vals
            else:
                key = i - (header_row + header_span - 1)  # positional 1-based
            self._key_to_phys[key] = i

    def _read(self, phys_row: int, phys_col: int):
        """Read a single cell via source — per-call open preserves live-read semantics."""
        return self.source.read_cell(self.sheet, phys_row, phys_col)

    def query(self, row, column) -> ExtractedValue:
        if column not in self._col_to_phys:
            raise KeyError(f"unknown column: {column!r}")
        if row not in self._key_to_phys:
            raise KeyError(f"unknown row key: {row!r}")
        pc, pr = self._col_to_phys[column], self._key_to_phys[row]
        spec = self.columns.get(column)
        return ExtractedValue(
            value=self._read(pr, pc),
            dtype=spec.dtype if spec else "string",
            unit=spec.unit if spec else None,
            sheet=self.sheet,
            cell_ref=f"{get_column_letter(pc)}{pr}",
            is_computed=bool(spec and spec.role == "computed"),
        )

    def query_cell(self, a1) -> ExtractedValue:
        from openpyxl.utils import coordinate_to_tuple
        r, c = coordinate_to_tuple(a1)
        return ExtractedValue(
            value=self._read(r, c),
            dtype="number",
            unit=None,
            sheet=self.sheet,
            cell_ref=a1,
            is_computed=False,
        )

    def query_range(self, a1) -> list[ExtractedValue]:
        min_row, min_col, max_row, max_col = range_box(a1)
        rows = self.source.read_region(self.sheet, min_row, min_col, max_row, max_col)
        out = []
        for r_off, row in enumerate(rows):
            r = min_row + r_off
            for c_off, val in enumerate(row):
                c = min_col + c_off
                out.append(ExtractedValue(
                    value=val,
                    dtype="number",
                    unit=None,
                    sheet=self.sheet,
                    cell_ref=f"{get_column_letter(c)}{r}",
                    is_computed=False,
                ))
        return out

    def physical_columns(self) -> dict:
        """Column name -> absolute 1-based physical column (copy; safe to mutate)."""
        return dict(self._col_to_phys)

    def data_row_numbers(self) -> list:
        """Sorted absolute 1-based physical row numbers of data rows."""
        return sorted(set(self._key_to_phys.values()))

    def read_all(self, max_rows: int | None = None) -> list[tuple]:
        """Open the workbook ONCE and return all (row_key, col_name, value, cell_ref) tuples.

        Respects ``_key_to_phys`` / ``_col_to_phys`` so values are consistent with
        ``query()``.  If *max_rows* is given, only the first *max_rows* row keys are
        read (order = insertion order of ``_key_to_phys``).

        Does NOT affect the live-read property of ``query()`` — each call here opens
        a fresh workbook handle and closes it when done.
        """
        row_keys = list(self._key_to_phys.keys())
        if max_rows is not None:
            row_keys = row_keys[:max_rows]

        col_items = list(self._col_to_phys.items())  # [(col_name, phys_col), ...]

        # OPT-2: one bounding-box read_region call (not per-cell) — per-cell opens
        # reintroduce the large-file hang. Index into the returned grid below.
        if not row_keys or not col_items:
            return []

        phys_rows = [self._key_to_phys[rk] for rk in row_keys]
        phys_cols = [pc for _, pc in col_items]
        min_row, max_row = min(phys_rows), max(phys_rows)
        min_col, max_col = min(phys_cols), max(phys_cols)
        grid = self.source.read_region(self.sheet, min_row, min_col, max_row, max_col)

        out: list[tuple] = []
        for row_key in row_keys:
            phys_row = self._key_to_phys[row_key]
            row = grid[phys_row - min_row]
            for col_name, phys_col in col_items:
                value = row[phys_col - min_col]
                cell_ref = f"{get_column_letter(phys_col)}{phys_row}"
                out.append((row_key, col_name, value, cell_ref))
        return out

    def row_keys(self) -> list:
        return list(self._key_to_phys.keys())

    def column_names(self) -> list[str]:
        return list(self._col_to_phys.keys())


def build_index(source, handle, row_key) -> ExtractionIndex:
    from mcg_swarm.source import as_source
    header_span = getattr(handle, "header_span", 1)
    return ExtractionIndex(
        as_source(source), handle.sheet, handle.region, handle.header_row, handle.columns,
        row_key, header_span=header_span,
    )
